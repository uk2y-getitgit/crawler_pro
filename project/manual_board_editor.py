# -*- coding: utf-8 -*-
"""
manual_board_editor.py — 게시판 수동 등록/수정 도구

수집 실패 또는 오탐된 사이트의 게시판 URL을 수동으로 등록한다.

사용법:
  python manual_board_editor.py            # 실패 목록 전체 편집
  python manual_board_editor.py 대전광역시  # 특정 기관만 수정
  python manual_board_editor.py --list     # 현재 등록 현황 조회
  python manual_board_editor.py --show-failed  # 실패 목록 조회
"""

import os, sys, json
import openpyxl
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

SITES_PATH  = os.path.join(BASE_DIR, "config", "sites.xlsx")
FAILED_PATH = os.path.join(BASE_DIR, "data", "board_finder_failed.json")
RESULT_PATH = os.path.join(BASE_DIR, "data", "board_finder_result.json")


# ─────────────────────────────────────────────
# 엑셀 읽기/쓰기
# ─────────────────────────────────────────────
COL = {"기관명":1,"사이트타입":2,"게시판명":3,"게시판URL":4,
       "페이지파라미터":5,"활성화":6,"비고":7}

def load_sites():
    wb = openpyxl.load_workbook(SITES_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    return wb, ws, rows, header

def show_current_list():
    """현재 sites.xlsx 등록 현황 출력"""
    wb, ws, rows, _ = load_sites()
    print(f"\n{'기관명':<20} {'게시판명':<20} {'게시판URL':<50} {'비고'}")
    print("-" * 110)
    for row in rows[1:]:
        agency    = row[COL["기관명"]-1]   or ""
        bname     = row[COL["게시판명"]-1]  or "❌ 미등록"
        burl      = row[COL["게시판URL"]-1] or ""
        note      = row[COL["비고"]-1]      or ""
        flag = "✅" if burl and bname != "❌ 미등록" else "❌"
        print(f"{flag} {agency:<18} {bname:<20} {burl[:48]:<50} {note}")

def show_failed_list():
    """실패 목록 출력"""
    if not os.path.exists(FAILED_PATH):
        print("실패 파일(data/board_finder_failed.json) 없음 — 먼저 수집을 실행하세요.")
        return
    with open(FAILED_PATH, encoding="utf-8") as f:
        data = json.load(f)
    failed = data.get("failed", {})
    if not failed:
        print("실패 목록이 없습니다. 모든 사이트 수집 성공!")
        return
    print(f"\n실패/미탐지 기관 목록 ({len(failed)}개):")
    print("-" * 60)
    for i, (agency, info) in enumerate(failed.items(), 1):
        note = info.get("note", "")
        print(f"  {i:>2}. {agency:<25} {note}")

def get_sites_row_for_agency(ws, rows, agency_name):
    """기관명으로 행 번호(1-based) 목록 반환"""
    matches = []
    for i, row in enumerate(rows[1:], 2):
        if row[COL["기관명"]-1] == agency_name:
            matches.append(i)
    return matches

def update_or_add_board(ws, rows, agency_name, board_name, board_url,
                        page_param="", note="수동등록", site_type=None, active="Y"):
    """
    게시판 정보를 sites.xlsx에 저장.
    기존 행에 게시판URL이 없으면 덮어씀.
    이미 채워진 행이 있으면 새 행 추가.
    """
    existing_rows = get_sites_row_for_agency(ws, rows, agency_name)

    # 기존 행 중 게시판URL 비어있는 것 찾기
    target_row = None
    for row_idx in existing_rows:
        row_data = rows[row_idx - 1]
        existing_url = row_data[COL["게시판URL"]-1]
        if not existing_url:
            target_row = row_idx
            break

    if target_row:
        ws.cell(target_row, COL["게시판명"]).value  = board_name
        ws.cell(target_row, COL["게시판URL"]).value  = board_url
        ws.cell(target_row, COL["페이지파라미터"]).value = page_param
        ws.cell(target_row, COL["비고"]).value       = note
        print(f"  ✅ {agency_name} [{board_name}] 기존 행 업데이트 (row {target_row})")
    else:
        # 새 행 추가: 마지막 기존 행 뒤에
        if existing_rows:
            last_row = existing_rows[-1]
            last_data = rows[last_row - 1]
            st = site_type or last_data[COL["사이트타입"]-1] or "일반"
        else:
            last_row = ws.max_row
            st = site_type or "일반"
        ws.insert_rows(last_row + 1)
        insert_row = last_row + 1
        ws.cell(insert_row, COL["기관명"]).value      = agency_name
        ws.cell(insert_row, COL["사이트타입"]).value  = st
        ws.cell(insert_row, COL["게시판명"]).value    = board_name
        ws.cell(insert_row, COL["게시판URL"]).value   = board_url
        ws.cell(insert_row, COL["페이지파라미터"]).value = page_param
        ws.cell(insert_row, COL["활성화"]).value      = active
        ws.cell(insert_row, COL["비고"]).value         = note
        print(f"  ✅ {agency_name} [{board_name}] 새 행 추가 (row {insert_row})")


# ─────────────────────────────────────────────
# 대화형 편집기
# ─────────────────────────────────────────────
def interactive_edit(agency_filter=None):
    wb, ws, rows, _ = load_sites()

    if agency_filter:
        agencies = agency_filter
        print(f"\n수동 편집 대상: {', '.join(agencies)}")
    else:
        # 실패 목록 우선
        if os.path.exists(FAILED_PATH):
            with open(FAILED_PATH, encoding="utf-8") as f:
                failed_data = json.load(f)
            agencies = list(failed_data.get("failed", {}).keys())
        else:
            # sites.xlsx에서 게시판URL 없는 기관
            agencies = [row[COL["기관명"]-1] for row in rows[1:]
                        if row[COL["기관명"]-1] and not row[COL["게시판URL"]-1]]
            agencies = list(dict.fromkeys(agencies))  # 중복 제거

    if not agencies:
        print("수동 등록이 필요한 기관이 없습니다.")
        return

    print(f"\n총 {len(agencies)}개 기관 수동 등록 시작")
    print("입력 방법: 게시판명|게시판URL|페이지파라미터(선택)")
    print("  예: 고시공고|https://www.xxx.go.kr/notice/list.do|?page=")
    print("  s 입력: 이 기관 건너뜀 | q 입력: 저장 후 종료\n")

    changed = 0
    for i, agency in enumerate(agencies, 1):
        print(f"\n[{i}/{len(agencies)}] {agency}")

        # 현재 등록된 게시판 보여주기
        cur_rows = get_sites_row_for_agency(ws, rows, agency)
        if cur_rows:
            for r in cur_rows:
                rd = rows[r-1]
                cur_bname = rd[COL["게시판명"]-1] or ""
                cur_burl  = rd[COL["게시판URL"]-1] or ""
                print(f"  현재: [{cur_bname}] {cur_burl or '(미등록)'}")

        user_input = input("  입력 (s=건너뜀, q=종료): ").strip()
        if not user_input or user_input.lower() == "s":
            continue
        if user_input.lower() == "q":
            break

        parts = [p.strip() for p in user_input.split("|")]
        if len(parts) < 2:
            print("  형식 오류 — '게시판명|URL' 형태로 입력하세요. 건너뜁니다.")
            continue

        board_name  = parts[0]
        board_url   = parts[1]
        page_param  = parts[2] if len(parts) > 2 else ""

        # 게시판 여러개 추가 (같은 기관)
        update_or_add_board(ws, rows, agency, board_name, board_url, page_param)
        changed += 1

        # 추가 게시판 여부
        while True:
            more = input("  이 기관에 게시판 추가? (y=추가, 엔터=다음 기관): ").strip().lower()
            if more != "y":
                break
            extra = input("  추가 입력 (게시판명|URL|페이지파라미터): ").strip()
            if not extra:
                break
            parts2 = [p.strip() for p in extra.split("|")]
            if len(parts2) >= 2:
                update_or_add_board(ws, rows, agency, parts2[0], parts2[1],
                                    parts2[2] if len(parts2) > 2 else "")
                changed += 1

    # 저장
    try:
        wb.save(SITES_PATH)
        print(f"\n✅ 저장 완료 (변경 {changed}건): {SITES_PATH}")
    except PermissionError:
        backup = SITES_PATH.replace(".xlsx", f"_manual_{datetime.now().strftime('%H%M%S')}.xlsx")
        wb.save(backup)
        print(f"\n⚠️  파일이 열려있어 백업으로 저장: {backup}")


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    args = sys.argv[1:]

    if "--list" in args:
        show_current_list()
        return

    if "--show-failed" in args:
        show_failed_list()
        return

    agencies = [a for a in args if not a.startswith("-")]
    interactive_edit(agency_filter=agencies if agencies else None)


if __name__ == "__main__":
    main()
