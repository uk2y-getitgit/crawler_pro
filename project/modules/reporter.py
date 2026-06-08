# -*- coding: utf-8 -*-
"""
reporter.py — 안전점검 모니터링 시스템 Phase 2

수집 결과를 엑셀(results/결과_YYYYMMDD.xlsx)로 저장한다.
3개 시트: 신규공고 / 전체공고 / 오류로그.
"""

import os
import sys
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from loguru import logger  # type: ignore
except ImportError:
    logger = logging.getLogger("reporter")
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        _h = logging.StreamHandler()
        _h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        logger.addHandler(_h)


if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


POST_HEADERS = [
    "신규여부", "수집일시", "기관명", "게시판명",
    "게시글제목", "게시일", "URL", "AI판단근거",
]
EXCLUDED_HEADERS = [
    "신규여부", "수집일시", "기관명", "게시판명",
    "게시글제목", "게시일", "URL", "제외사유",
]
ERROR_HEADERS = ["기관명", "게시판URL", "오류내용", "재시도횟수", "처리결과"]

YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99",
                          fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9",
                          fill_type="solid")
HEADER_FONT = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = [10, 18, 16, 14, 50, 12, 45, 30]


class Reporter:
    def __init__(self, results_dir="results", base_dir=None):
        self.base_dir = base_dir or BASE_DIR
        if os.path.isabs(results_dir):
            self.results_dir = results_dir
        else:
            self.results_dir = os.path.join(self.base_dir, results_dir)
        os.makedirs(self.results_dir, exist_ok=True)

    def save(self, results, errors=None, filename=None, excluded=None):
        errors = errors or []
        excluded = excluded or []
        now = datetime.now()
        collected_at = now.strftime("%Y-%m-%d %H:%M")
        if filename is None:
            filename = f"결과_{now.strftime('%Y%m%d')}.xlsx"
        out_path = os.path.join(self.results_dir, filename)

        wb = openpyxl.Workbook()

        # 1) 신규공고
        ws_new = wb.active
        ws_new.title = "신규공고"
        new_items = [r for r in results if r.get("is_new")]
        self._write_posts_sheet(ws_new, new_items, collected_at, highlight=True)

        # 2) 전체공고 (최신순 정렬)
        ws_all = wb.create_sheet("전체공고")
        all_sorted = sorted(
            results,
            key=lambda r: (r.get("date") or ""),
            reverse=True,
        )
        self._write_posts_sheet(ws_all, all_sorted, collected_at, highlight=False)

        # 3) 제외공고 (AI 제외 + 게시일 초과) — 통과/제외 추적용
        ws_ex = wb.create_sheet("제외공고")
        ex_sorted = sorted(excluded, key=lambda r: (r.get("date") or ""),
                           reverse=True)
        self._write_posts_sheet(ws_ex, ex_sorted, collected_at,
                                highlight=False, headers=EXCLUDED_HEADERS)

        # 4) 오류로그
        ws_err = wb.create_sheet("오류로그")
        self._write_error_sheet(ws_err, errors)

        wb.save(out_path)
        logger.info(f"엑셀 저장: {out_path} "
                    f"(신규 {len(new_items)} / 전체 {len(results)} / "
                    f"제외 {len(excluded)} / 오류 {len(errors)})")
        return out_path

    # ------------------------------------------------------------- internals
    def _write_header(self, ws, headers):
        for col, name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

    def _write_posts_sheet(self, ws, items, collected_at, highlight,
                           headers=None):
        self._write_header(ws, headers or POST_HEADERS)
        for r, item in enumerate(items, start=2):
            is_new = item.get("is_new")
            values = [
                "신규" if is_new else "기존",
                collected_at,
                item.get("agency", ""),
                item.get("board_name", ""),
                item.get("title", ""),
                item.get("date", ""),
                item.get("url", ""),
                item.get("ai_reason", ""),
            ]
            for c, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if highlight or is_new:
                    cell.fill = YELLOW_FILL
        self._apply_widths(ws, COL_WIDTHS)
        ws.freeze_panes = "A2"
        # URL 하이퍼링크
        for r in range(2, ws.max_row + 1):
            url_cell = ws.cell(row=r, column=7)
            if url_cell.value:
                url_cell.hyperlink = url_cell.value
                url_cell.font = Font(color="0563C1", underline="single")

    def _write_error_sheet(self, ws, errors):
        self._write_header(ws, ERROR_HEADERS)
        for r, e in enumerate(errors, start=2):
            values = [
                e.get("agency", ""),
                e.get("url", ""),
                e.get("error", ""),
                e.get("retries", ""),
                e.get("outcome", ""),
            ]
            for c, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")
        self._apply_widths(ws, [16, 45, 40, 12, 12])
        ws.freeze_panes = "A2"

    def _apply_widths(self, ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
