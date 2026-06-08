# -*- coding: utf-8 -*-
"""
gui_dialogs.py — 안전점검 모니터링 시스템 Phase 5

- SettingsDialog: AI_PROVIDER 전환, API 키 마스킹 표시, 크롤 딜레이/스케줄 설정.
- BoardFinderThread: BoardFinder를 백그라운드로 실행하는 QThread.
- BoardFinderDialog: 기관 선택 후 게시판 자동 탐색 실행.
- BoardManageDialog: 수집 결과 조회 + 실패/오탐 사이트 수동 등록/수정.

설정은 .env 파일에 읽기/쓰기한다(프로젝트 전반이 .env 기반).
"""

from __future__ import annotations

import json
import os
import logging
from datetime import datetime

from PyQt6.QtCore import QThread, pyqtSignal, Qt
from PyQt6.QtGui import QColor, QBrush
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QLabel, QLineEdit,
    QComboBox, QSpinBox, QCheckBox, QPushButton, QPlainTextEdit, QListWidget,
    QListWidgetItem, QDialogButtonBox, QTimeEdit, QGroupBox, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QSplitter,
    QWidget, QFileDialog,
)
from PyQt6.QtCore import QTime

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

try:
    from modules.board_finder import BoardFinder
except ImportError:  # 단독 실행/패키징 호환
    from board_finder import BoardFinder

try:
    from loguru import logger  # type: ignore
except ImportError:
    logger = logging.getLogger("gui_dialogs")


# --------------------------------------------------------------------------- #
# .env 읽기/쓰기 헬퍼
# --------------------------------------------------------------------------- #
def _read_env(env_path):
    """.env를 dict로 읽는다(순서 보존은 하지 않음)."""
    data = {}
    if not os.path.exists(env_path):
        return data
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                data[k.strip()] = v.strip()
    except OSError as e:
        logger.warning(f".env 읽기 실패: {e}")
    return data


def _write_env(env_path, updates):
    """.env에 updates(dict)를 병합 저장한다. 기존 키는 갱신, 없으면 추가."""
    existing = _read_env(env_path)
    existing.update({k: str(v) for k, v in updates.items()})
    try:
        tmp = env_path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            for k, v in existing.items():
                f.write(f"{k}={v}\n")
        os.replace(tmp, env_path)  # 원자적 교체 — 쓰기 중 손상 방지
        return True
    except OSError as e:
        logger.warning(f".env 저장 실패: {e}")
        return False


def _mask_key(key):
    """API 키를 마스킹한다. 앞 7자 + **** + 뒤 4자."""
    if not key:
        return ""
    key = key.strip()
    if len(key) <= 12:
        return "*" * len(key)
    return f"{key[:7]}{'*' * 8}{key[-4:]}"


# --------------------------------------------------------------------------- #
# SettingsDialog
# --------------------------------------------------------------------------- #
class SettingsDialog(QDialog):
    """AI provider / API 키 / 크롤 딜레이 / 스케줄 설정."""

    def __init__(self, base_dir, parent=None):
        super().__init__(parent)
        self.base_dir = base_dir
        self.env_path = os.path.join(base_dir, ".env")
        self._env = _read_env(self.env_path)

        self.setWindowTitle("설정")
        self.setMinimumWidth(460)
        self._build_ui()
        self._load_values()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # --- AI 설정 ---
        ai_box = QGroupBox("AI 설정")
        ai_form = QFormLayout(ai_box)

        self.provider_combo = QComboBox()
        self.provider_combo.addItem("Claude", "claude")
        self.provider_combo.addItem("Gemini", "gemini")
        ai_form.addRow("AI Provider", self.provider_combo)

        self.claude_key_label = QLabel("-")
        self.claude_key_label.setTextInteractionFlags(
            Qt.TextInteractionFlag.TextSelectableByMouse)
        ai_form.addRow("Claude API 키", self.claude_key_label)

        self.gemini_key_label = QLabel("-")
        self.gemini_key_label.setTextInteractionFlags(
            Qt.TextInteractionFlag.TextSelectableByMouse)
        ai_form.addRow("Gemini API 키", self.gemini_key_label)

        # 키 편집(선택) — 비워두면 기존 값 유지
        self.claude_key_edit = QLineEdit()
        self.claude_key_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.claude_key_edit.setPlaceholderText("새 Claude 키 입력 시에만(비우면 유지)")
        ai_form.addRow("Claude 키 변경", self.claude_key_edit)

        self.gemini_key_edit = QLineEdit()
        self.gemini_key_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.gemini_key_edit.setPlaceholderText("새 Gemini 키 입력 시에만(비우면 유지)")
        ai_form.addRow("Gemini 키 변경", self.gemini_key_edit)

        layout.addWidget(ai_box)

        # --- 크롤 설정 ---
        crawl_box = QGroupBox("크롤 설정")
        crawl_form = QFormLayout(crawl_box)
        self.delay_spin = QSpinBox()
        self.delay_spin.setRange(0, 60)
        self.delay_spin.setSuffix(" 초")
        crawl_form.addRow("사이트 간 딜레이", self.delay_spin)

        self.maxpages_spin = QSpinBox()
        self.maxpages_spin.setRange(1, 20)
        crawl_form.addRow("최대 페이지 수", self.maxpages_spin)

        self.searchdays_spin = QSpinBox()
        self.searchdays_spin.setRange(0, 365)
        self.searchdays_spin.setSuffix(" 일")
        self.searchdays_spin.setToolTip("게시일이 최근 N일 이내인 공고만 통과 "
                                        "(0 = 날짜 제한 없음)")
        crawl_form.addRow("게시일 기간", self.searchdays_spin)
        layout.addWidget(crawl_box)

        # --- 스케줄 설정 ---
        sched_box = QGroupBox("자동 실행 스케줄")
        sched_form = QFormLayout(sched_box)
        self.sched_enabled = QCheckBox("매일 지정 시각에 자동 크롤링")
        sched_form.addRow(self.sched_enabled)
        self.sched_time = QTimeEdit()
        self.sched_time.setDisplayFormat("HH:mm")
        sched_form.addRow("실행 시각", self.sched_time)
        layout.addWidget(sched_box)

        # --- 버튼 ---
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save
            | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._on_save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _load_values(self):
        provider = (self._env.get("AI_PROVIDER", "claude") or "claude").lower()
        idx = self.provider_combo.findData(provider)
        self.provider_combo.setCurrentIndex(idx if idx >= 0 else 0)

        self.claude_key_label.setText(
            _mask_key(self._env.get("ANTHROPIC_API_KEY", "")) or "(미설정)")
        self.gemini_key_label.setText(
            _mask_key(self._env.get("GEMINI_API_KEY", "")) or "(미설정)")

        try:
            self.delay_spin.setValue(int(self._env.get("CRAWL_DELAY", 2)))
        except (TypeError, ValueError):
            self.delay_spin.setValue(2)
        try:
            self.maxpages_spin.setValue(int(self._env.get("MAX_PAGES", 3)))
        except (TypeError, ValueError):
            self.maxpages_spin.setValue(3)
        try:
            self.searchdays_spin.setValue(int(self._env.get("SEARCH_DAYS", 3)))
        except (TypeError, ValueError):
            self.searchdays_spin.setValue(3)

        self.sched_enabled.setChecked(
            str(self._env.get("SCHEDULE_ENABLED", "N")).upper() == "Y")
        t = self._env.get("SCHEDULE_TIME", "09:00") or "09:00"
        qt = QTime.fromString(t, "HH:mm")
        if qt.isValid():
            self.sched_time.setTime(qt)
        else:
            self.sched_time.setTime(QTime(9, 0))

    def _on_save(self):
        updates = {
            "AI_PROVIDER": self.provider_combo.currentData(),
            "CRAWL_DELAY": self.delay_spin.value(),
            "MAX_PAGES": self.maxpages_spin.value(),
            "SEARCH_DAYS": self.searchdays_spin.value(),
            "SCHEDULE_ENABLED": "Y" if self.sched_enabled.isChecked() else "N",
            "SCHEDULE_TIME": self.sched_time.time().toString("HH:mm"),
        }
        # 키는 입력된 경우에만 갱신(마스킹된 값을 다시 쓰지 않음)
        new_claude = self.claude_key_edit.text().strip()
        if new_claude:
            updates["ANTHROPIC_API_KEY"] = new_claude
        new_gemini = self.gemini_key_edit.text().strip()
        if new_gemini:
            updates["GEMINI_API_KEY"] = new_gemini

        if _write_env(self.env_path, updates):
            # 현재 프로세스 환경에도 즉시 반영
            for k, v in updates.items():
                os.environ[k] = str(v)
            QMessageBox.information(self, "설정", "설정을 저장했습니다.")
            self.accept()
        else:
            QMessageBox.warning(self, "설정", ".env 저장에 실패했습니다.")


# --------------------------------------------------------------------------- #
# BoardFinderThread
# --------------------------------------------------------------------------- #
class BoardFinderThread(QThread):
    """BoardFinder.run_all + save_to_sites를 백그라운드로 실행."""

    progress = pyqtSignal(str, dict)
    result_ready = pyqtSignal(dict)   # {기관명: [{name,url}]}
    error = pyqtSignal(str)

    def __init__(self, sites_path, agency_filter=None, base_dir=None,
                 auto_save=True, parent=None):
        super().__init__(parent)
        self.sites_path = sites_path
        self.agency_filter = agency_filter
        self.base_dir = base_dir
        self.auto_save = auto_save

    def run(self):
        try:
            finder = BoardFinder(
                sites_path=self.sites_path,
                progress_callback=self._on_progress,
                base_dir=self.base_dir,
            )
            results = finder.run_all(agency_filter=self.agency_filter)
            if self.auto_save:
                try:
                    finder.save_to_sites(results)
                except PermissionError:
                    self.error.emit(
                        "sites.xlsx가 열려 있어 저장하지 못했습니다. "
                        "엑셀에서 파일을 닫고 다시 시도하세요.")
                    return
            self.result_ready.emit(results)
        except Exception as e:
            logger.exception("BoardFinderThread 실행 중 치명 오류")
            self.error.emit(str(e))

    def _on_progress(self, event_type, data):
        if not isinstance(data, dict):
            data = {"value": data}
        self.progress.emit(str(event_type), data)


# --------------------------------------------------------------------------- #
# BoardFinderDialog
# --------------------------------------------------------------------------- #
class BoardFinderDialog(QDialog):
    """기관 선택 후 게시판 자동 탐색을 실행하는 다이얼로그."""

    def __init__(self, base_dir, sites_path, parent=None):
        super().__init__(parent)
        self.base_dir = base_dir
        self.sites_path = sites_path
        self._thread = None

        self.setWindowTitle("게시판 등록 (자동 탐색)")
        self.setMinimumSize(560, 480)
        self._build_ui()
        self._load_agencies()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            "게시판을 탐색할 기관을 선택하세요. (선택 없으면 전체 실행)"))

        self.agency_list = QListWidget()
        self.agency_list.setSelectionMode(
            QListWidget.SelectionMode.MultiSelection)
        layout.addWidget(self.agency_list)

        btn_row = QHBoxLayout()
        self.run_btn = QPushButton("탐색 실행")
        self.run_btn.clicked.connect(self._on_run)
        self.stop_btn = QPushButton("닫기")
        self.stop_btn.clicked.connect(self.reject)
        btn_row.addWidget(self.run_btn)
        btn_row.addStretch(1)
        btn_row.addWidget(self.stop_btn)
        layout.addLayout(btn_row)

        layout.addWidget(QLabel("진행 로그"))
        self.log = QPlainTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)

    def _load_agencies(self):
        """sites.xlsx에서 기관명 목록을 읽어 채운다."""
        if openpyxl is None:
            self._append_log("openpyxl 미설치 — 기관 목록을 읽을 수 없습니다.")
            return
        path = self.sites_path
        if not os.path.isabs(path):
            path = os.path.join(self.base_dir, path)
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            seen = set()
            for row in list(ws.iter_rows(values_only=True))[1:]:
                if not row or row[0] is None:
                    continue
                agency = str(row[0]).strip()
                if agency and agency not in seen:
                    seen.add(agency)
                    self.agency_list.addItem(QListWidgetItem(agency))
            wb.close()
        except Exception as e:
            self._append_log(f"기관 목록 로드 실패: {e}")

    def _selected_agencies(self):
        return [i.text() for i in self.agency_list.selectedItems()]

    def _on_run(self):
        if self._thread is not None and self._thread.isRunning():
            return
        agencies = self._selected_agencies()
        agency_filter = agencies if agencies else None
        scope = ", ".join(agencies) if agencies else "전체 사이트"
        self._append_log(f"게시판 탐색 시작: {scope}")
        self.run_btn.setEnabled(False)

        self._thread = BoardFinderThread(
            sites_path=self.sites_path,
            agency_filter=agency_filter,
            base_dir=self.base_dir,
            auto_save=True,
        )
        self._thread.progress.connect(self._on_progress)
        self._thread.result_ready.connect(self._on_done)
        self._thread.error.connect(self._on_error)
        self._thread.finished.connect(lambda: self.run_btn.setEnabled(True))
        self._thread.start()

    def _on_progress(self, event_type, data):
        self._append_log(f"[{event_type}] {data}")

    def _on_done(self, results):
        total = sum(len(v) for v in results.values())
        self._append_log(
            f"완료: 기관 {len(results)}개, 게시판 {total}개 선별 — sites.xlsx 저장됨")
        QMessageBox.information(self, "게시판 등록",
                                f"게시판 {total}개를 sites.xlsx에 저장했습니다.")

    def _on_error(self, msg):
        self._append_log(f"[오류] {msg}")
        QMessageBox.warning(self, "게시판 등록", msg)

    def _append_log(self, text):
        self.log.appendPlainText(text)

    def closeEvent(self, event):
        """다이얼로그 종료 시 스레드가 끝나길 기다린다(자원 누수 방지)."""
        if self._thread is not None and self._thread.isRunning():
            self._thread.wait(5000)
        super().closeEvent(event)


# --------------------------------------------------------------------------- #
# BoardManageDialog — 게시판 관리 (조회 / 수동 등록·수정·삭제 / 엑셀 일괄가져오기)
# --------------------------------------------------------------------------- #
_COL = {"기관명": 0, "사이트타입": 1, "게시판명": 2, "게시판URL": 3,
        "페이지파라미터": 4, "활성화": 5, "비고": 6}

COLOR_OK   = QColor("#C8E6C9")   # 연초록 — 게시판 등록됨
COLOR_FAIL = QColor("#FFCDD2")   # 연빨강 — 미등록
COLOR_PART = QColor("#FFF9C4")   # 연노랑 — 일부만 등록

# 게시판 행(활성화 상태) 색상
COLOR_ACTIVE   = QColor("#FFFFFF")   # 흰색 — 크롤링 가능(활성 Y)
COLOR_MANUAL   = QColor("#FFF9C4")   # 노랑 — 수동검색 필요(활성 N)
COLOR_INACTIVE = QColor("#FFCDD2")   # 빨강 — 크롤링 불가(활성 N)

# 수동검색으로 판단하는 비고 키워드
_MANUAL_NOTE_KW = ("수동", "조회버튼", "헤더메뉴", "키워드 검색", "검색 필요",
                   "개별 게시판", "나라장터")


def _row_color_for(active, note):
    """활성화/비고로 게시판 행 색상 결정. (가능=흰, 수동=노랑, 불가=빨강)"""
    act = str(active or "Y").strip().upper()
    note = str(note or "")
    if act == "Y":
        return COLOR_ACTIVE
    if "로그인우회" not in note and any(k in note for k in _MANUAL_NOTE_KW):
        return COLOR_MANUAL
    return COLOR_INACTIVE


# ─────────────────────────────────────────────
# 게시판 1건 추가/수정 팝업
# ─────────────────────────────────────────────
class _BoardEditDialog(QDialog):
    def __init__(self, parent=None, name="", url="", param=""):
        super().__init__(parent)
        self.setWindowTitle("게시판 정보 입력")
        self.setMinimumWidth(500)
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        form = QFormLayout()
        self.name_edit  = QLineEdit(name)
        self.name_edit.setPlaceholderText("예: 입찰공고")
        self.url_edit   = QLineEdit(url)
        self.url_edit.setPlaceholderText("https://...")
        self.param_edit = QLineEdit(param)
        self.param_edit.setPlaceholderText("?page=  (페이지 파라미터, 없으면 빈칸)")
        form.addRow("게시판명  *", self.name_edit)
        form.addRow("게시판 URL *", self.url_edit)
        form.addRow("페이지 파라미터", self.param_edit)
        layout.addLayout(form)

        hint = QLabel("※ URL은 게시판 목록 첫 페이지 주소를 입력하세요.")
        hint.setStyleSheet("color:#666; font-size:11px;")
        layout.addWidget(hint)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self._on_ok)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _on_ok(self):
        if not self.name_edit.text().strip():
            QMessageBox.warning(self, "입력 오류", "게시판명을 입력하세요.")
            return
        if not self.url_edit.text().strip():
            QMessageBox.warning(self, "입력 오류", "게시판 URL을 입력하세요.")
            return
        self.accept()

    def values(self):
        return (self.name_edit.text().strip(),
                self.url_edit.text().strip(),
                self.param_edit.text().strip())


# ─────────────────────────────────────────────
# 메인 게시판 관리 다이얼로그
# ─────────────────────────────────────────────
class BoardManageDialog(QDialog):
    """
    게시판 현황 조회 + 수동 등록/수정/삭제 + 엑셀 일괄가져오기.

    왼쪽: 기관 목록 (초록=등록됨, 빨강=미등록, 노랑=일부)
    오른쪽: 게시판 테이블 + 추가·수정·삭제 버튼
    하단: 저장 / 엑셀 일괄가져오기 / 템플릿 다운로드 / 닫기
    """

    def __init__(self, base_dir, sites_path, parent=None):
        super().__init__(parent)
        self.base_dir   = base_dir
        self.sites_path = sites_path
        self._wb        = None
        self._ws        = None
        self._dirty     = False
        self._agency_rows: dict = {}

        self.setWindowTitle("게시판 관리")
        self.setMinimumSize(960, 620)
        self._build_ui()
        self._load_data()

    # ─────────────── UI 구성
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)

        # ── 상단 타이틀 + 범례 ──
        top_bar = QHBoxLayout()
        top_bar.setContentsMargins(0, 0, 0, 0)
        top_bar.setSpacing(6)
        title = QLabel("게시판 관리")
        title.setStyleSheet("font-size:15px; font-weight:bold; color:#1a3c6e;")
        top_bar.addWidget(title)
        top_bar.addStretch()
        legend_lbl = QLabel("범례:")
        legend_lbl.setStyleSheet("font-size:11px; color:#666;")
        legend_lbl.setFixedHeight(22)
        top_bar.addWidget(legend_lbl)
        for txt, bg in [("등록됨", "#C8E6C9"),
                        ("일부등록", "#FFF9C4"),
                        ("미등록", "#FFCDD2")]:
            lbl = QLabel(txt)
            lbl.setFixedHeight(22)
            lbl.setStyleSheet(
                f"background:{bg}; border:1px solid #bbb; border-radius:3px;"
                " padding:0px 7px; font-size:11px;")
            lbl.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignHCenter)
            top_bar.addWidget(lbl)
        top_bar_widget = QWidget()
        top_bar_widget.setLayout(top_bar)
        top_bar_widget.setFixedHeight(30)
        root.addWidget(top_bar_widget)

        # ── 본문 분할 ──
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # ── 왼쪽 기관 목록 ──
        left = QWidget()
        ll   = QVBoxLayout(left)
        ll.setContentsMargins(0, 0, 0, 0)
        ll.setSpacing(4)

        self.agency_search = QLineEdit()
        self.agency_search.setPlaceholderText("기관명 검색…")
        self.agency_search.textChanged.connect(self._filter_agency_list)
        ll.addWidget(self.agency_search)

        self.agency_list = QListWidget()
        self.agency_list.setAlternatingRowColors(True)
        self.agency_list.currentRowChanged.connect(self._on_agency_select)
        ll.addWidget(self.agency_list)

        self.lbl_agency_stat = QLabel()
        self.lbl_agency_stat.setStyleSheet("font-size:11px; color:#555; padding:2px;")
        ll.addWidget(self.lbl_agency_stat)
        left.setMinimumWidth(230)

        # ── 오른쪽 게시판 ──
        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(6)

        self.agency_label = QLabel("기관을 선택하세요")
        self.agency_label.setStyleSheet(
            "font-size:14px; font-weight:bold; color:#1a3c6e; padding:4px 0;")
        rl.addWidget(self.agency_label)

        self.board_table = QTableWidget(0, 4)
        self.board_table.setHorizontalHeaderLabels(
            ["게시판명", "URL", "페이지파라미터", "활성화"])
        self.board_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.board_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.board_table.setAlternatingRowColors(False)  # 행 색상(활성화)이 보이도록
        self.board_table.verticalHeader().setVisible(False)
        self.board_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self.board_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.board_table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.ResizeMode.ResizeToContents)
        self.board_table.itemDoubleClicked.connect(lambda _: self._on_edit())
        rl.addWidget(self.board_table)

        # 버튼 행
        btn_row = QHBoxLayout()
        self.btn_add  = QPushButton("+ 게시판 추가")
        self.btn_edit = QPushButton("수정")
        self.btn_del  = QPushButton("삭제")
        self.btn_active = QPushButton("활성화 ON/OFF")
        self.btn_refind = QPushButton("AI 자동재탐색")
        for btn, color in [(self.btn_add, "#2E7D32"),
                           (self.btn_edit, "#1565C0"),
                           (self.btn_del,  "#C62828"),
                           (self.btn_active, "#00838F"),
                           (self.btn_refind, "#6A1B9A")]:
            btn.setStyleSheet(
                f"background:{color}; color:#fff; font-weight:600;"
                " border-radius:5px; padding:6px 16px;")
        self.btn_add.clicked.connect(self._on_add)
        self.btn_edit.clicked.connect(self._on_edit)
        self.btn_del.clicked.connect(self._on_delete)
        self.btn_active.clicked.connect(self._on_toggle_active)
        self.btn_refind.clicked.connect(self._on_refind)
        btn_row.addWidget(self.btn_add)
        btn_row.addWidget(self.btn_edit)
        btn_row.addWidget(self.btn_del)
        btn_row.addWidget(self.btn_active)
        btn_row.addStretch()
        btn_row.addWidget(self.btn_refind)
        rl.addLayout(btn_row)

        hint = QLabel("더블클릭으로 수정  |  Del 키로 삭제  |  "
                      "행 색상 — 흰색:크롤링 가능 / 노랑:수동검색 필요 / 빨강:크롤링 불가(비활성)")
        hint.setStyleSheet("font-size:11px; color:#888;")
        rl.addWidget(hint)

        splitter.addWidget(left)
        splitter.addWidget(right)
        splitter.setSizes([240, 680])
        root.addWidget(splitter)

        # ── 하단 버튼 바 ──
        bottom = QHBoxLayout()
        self.btn_save = QPushButton("저장")
        self.btn_save.setStyleSheet(
            "background:#1a3c6e; color:#fff; font-weight:bold;"
            " border-radius:5px; padding:7px 22px;")
        self.btn_save.clicked.connect(self._on_save)

        btn_import = QPushButton("엑셀로 일괄가져오기")
        btn_import.setStyleSheet(
            "background:#E65100; color:#fff; border-radius:5px; padding:7px 16px;")
        btn_import.clicked.connect(self._on_bulk_import)

        btn_template = QPushButton("입력 템플릿 다운로드")
        btn_template.setStyleSheet(
            "background:#fff; color:#1a3c6e; border:1.5px solid #1a3c6e;"
            " border-radius:5px; padding:7px 14px;")
        btn_template.clicked.connect(self._on_download_template)

        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet(
            "background:#616161; color:#fff; border-radius:5px; padding:7px 16px;")
        btn_close.clicked.connect(self.accept)

        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color:#1a3c6e; font-weight:bold; font-size:12px;")

        bottom.addWidget(self.btn_save)
        bottom.addWidget(btn_import)
        bottom.addWidget(btn_template)
        bottom.addStretch()
        bottom.addWidget(self.status_label)
        bottom.addWidget(btn_close)
        root.addLayout(bottom)

    # ─────────────── 데이터 로드
    def _load_data(self):
        if openpyxl is None:
            QMessageBox.warning(self, "오류", "openpyxl이 설치되지 않았습니다.")
            return
        if not os.path.exists(self.sites_path):
            QMessageBox.warning(self, "오류",
                                f"sites.xlsx를 찾을 수 없습니다:\n{self.sites_path}")
            return
        try:
            self._wb = openpyxl.load_workbook(self.sites_path)
            self._ws = self._wb.active
        except Exception as e:
            QMessageBox.warning(self, "오류", f"sites.xlsx 열기 실패: {e}")
            return

        self._agency_rows = {}
        for i, row in enumerate(list(self._ws.iter_rows(values_only=True))[1:], 2):
            if not row or row[0] is None:
                continue
            agency = str(row[0]).strip()
            if agency:
                self._agency_rows.setdefault(agency, []).append(i)

        self._rebuild_agency_list()

    def _rebuild_agency_list(self, keyword=""):
        self.agency_list.clear()
        ok = fail = part = 0
        for agency, rows in self._agency_rows.items():
            if keyword and keyword.lower() not in agency.lower():
                continue
            has_board = any(self._ws.cell(r, _COL["게시판명"]+1).value for r in rows)
            all_have  = all(self._ws.cell(r, _COL["게시판명"]+1).value for r in rows)
            cnt = sum(1 for r in rows if self._ws.cell(r, _COL["게시판명"]+1).value)
            item = QListWidgetItem(f"{agency}  ({cnt})")
            item.setData(Qt.ItemDataRole.UserRole, agency)
            if has_board and all_have:
                item.setBackground(QBrush(COLOR_OK))
                ok += 1
            elif has_board:
                item.setBackground(QBrush(COLOR_PART))
                part += 1
            else:
                item.setBackground(QBrush(COLOR_FAIL))
                fail += 1
            self.agency_list.addItem(item)
        total = ok + fail + part
        self.lbl_agency_stat.setText(
            f"전체 {total}  등록 {ok}  일부 {part}  미등록 {fail}")

    def _filter_agency_list(self, text):
        self._rebuild_agency_list(keyword=text)

    # ─────────────── 기관 선택
    def _on_agency_select(self, row_idx):
        if row_idx < 0:
            return
        item = self.agency_list.item(row_idx)
        if not item:
            return
        agency = item.data(Qt.ItemDataRole.UserRole)
        self.agency_label.setText(f"【 {agency} 】 게시판 목록")
        self._load_board_table(agency)

    def _load_board_table(self, agency):
        self.board_table.setRowCount(0)
        for r in self._agency_rows.get(agency, []):
            name  = self._ws.cell(r, _COL["게시판명"]+1).value or ""
            url   = self._ws.cell(r, _COL["게시판URL"]+1).value or ""
            param = self._ws.cell(r, _COL["페이지파라미터"]+1).value or ""
            active = str(self._ws.cell(r, _COL["활성화"]+1).value or "Y").strip().upper()
            note  = self._ws.cell(r, _COL["비고"]+1).value or ""
            if not name and not url:
                continue
            tr = self.board_table.rowCount()
            self.board_table.insertRow(tr)
            self.board_table.setItem(tr, 0, QTableWidgetItem(str(name)))
            self.board_table.setItem(tr, 1, QTableWidgetItem(str(url)))
            self.board_table.setItem(tr, 2, QTableWidgetItem(str(param)))
            act_item = QTableWidgetItem("ON" if active == "Y" else "OFF")
            act_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.board_table.setItem(tr, 3, act_item)
            self.board_table.item(tr, 0).setData(Qt.ItemDataRole.UserRole, r)
            # 행 색상: 활성화/비고 기준 (흰=가능, 노랑=수동, 빨강=불가)
            color = _row_color_for(active, note)
            for c in range(4):
                self.board_table.item(tr, c).setBackground(QBrush(color))

    def _current_agency(self):
        item = self.agency_list.currentItem()
        return item.data(Qt.ItemDataRole.UserRole) if item else None

    # ─────────────── CRUD
    def _on_add(self):
        agency = self._current_agency()
        if not agency:
            QMessageBox.information(self, "게시판 추가", "기관을 먼저 선택하세요.")
            return
        dlg = _BoardEditDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        name, url, param = dlg.values()
        self._write_board(agency, name, url, param, note="수동등록")
        self._dirty = True
        self._refresh_agency_color(agency)
        self._load_board_table(agency)
        self._update_status()

    def _on_edit(self):
        agency  = self._current_agency()
        sel_row = self.board_table.currentRow()
        if not agency or sel_row < 0:
            QMessageBox.information(self, "수정", "수정할 게시판을 선택하세요.")
            return
        name_now  = self.board_table.item(sel_row, 0).text()
        url_now   = self.board_table.item(sel_row, 1).text()
        param_now = self.board_table.item(sel_row, 2).text()
        xlsx_row  = self.board_table.item(sel_row, 0).data(Qt.ItemDataRole.UserRole)
        dlg = _BoardEditDialog(self, name=name_now, url=url_now, param=param_now)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        name, url, param = dlg.values()
        self._ws.cell(xlsx_row, _COL["게시판명"]+1).value    = name
        self._ws.cell(xlsx_row, _COL["게시판URL"]+1).value   = url
        self._ws.cell(xlsx_row, _COL["페이지파라미터"]+1).value = param
        self._ws.cell(xlsx_row, _COL["비고"]+1).value        = "수동수정"
        self._dirty = True
        self._load_board_table(agency)
        self._update_status()

    def _on_delete(self):
        agency  = self._current_agency()
        sel_row = self.board_table.currentRow()
        if not agency or sel_row < 0:
            QMessageBox.information(self, "삭제", "삭제할 게시판을 선택하세요.")
            return
        name_now = self.board_table.item(sel_row, 0).text()
        xlsx_row = self.board_table.item(sel_row, 0).data(Qt.ItemDataRole.UserRole)
        if QMessageBox.question(
                self, "삭제 확인",
                f"[{name_now}] 게시판을 삭제하시겠습니까?") \
                != QMessageBox.StandardButton.Yes:
            return
        self._ws.cell(xlsx_row, _COL["게시판명"]+1).value    = None
        self._ws.cell(xlsx_row, _COL["게시판URL"]+1).value   = None
        self._ws.cell(xlsx_row, _COL["페이지파라미터"]+1).value = None
        self._ws.cell(xlsx_row, _COL["비고"]+1).value        = "수동삭제"
        self._dirty = True
        self._refresh_agency_color(agency)
        self._load_board_table(agency)
        self._update_status()

    def _on_toggle_active(self):
        """선택 게시판의 활성화(Y/N)를 토글한다. 저장 시 sites.xlsx에 반영되어
        다음 크롤링부터 자동 적용된다(활성화=Y만 수집)."""
        agency  = self._current_agency()
        sel_row = self.board_table.currentRow()
        if not agency or sel_row < 0:
            QMessageBox.information(self, "활성화", "활성/비활성 전환할 게시판을 선택하세요.")
            return
        xlsx_row = self.board_table.item(sel_row, 0).data(Qt.ItemDataRole.UserRole)
        cur = str(self._ws.cell(xlsx_row, _COL["활성화"]+1).value or "Y").strip().upper()
        new = "N" if cur == "Y" else "Y"
        self._ws.cell(xlsx_row, _COL["활성화"]+1).value = new
        self._apply_row_fill(xlsx_row, new)
        self._dirty = True
        self._load_board_table(agency)
        self.board_table.selectRow(sel_row)
        self._update_status()

    def _apply_row_fill(self, xlsx_row, active):
        """엑셀 행(1~7열) 배경색을 활성화/비고에 맞춰 칠한다(파일에도 반영)."""
        try:
            from openpyxl.styles import PatternFill
        except ImportError:
            return
        note = self._ws.cell(xlsx_row, _COL["비고"]+1).value
        color = _row_color_for(active, note)
        hexrgb = color.name().replace("#", "").upper()
        fill = PatternFill(fill_type=None) if hexrgb in ("FFFFFF", "FFF") \
            else PatternFill("solid", fgColor=hexrgb)
        for c in range(1, 8):
            self._ws.cell(xlsx_row, c).fill = fill

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Delete:
            self._on_delete()
        else:
            super().keyPressEvent(event)

    # ─────────────── AI 재탐색
    def _on_refind(self):
        agency = self._current_agency()
        if not agency:
            QMessageBox.information(self, "재탐색", "기관을 먼저 선택하세요.")
            return
        if QMessageBox.question(
                self, "AI 자동재탐색",
                f"[{agency}] 게시판을 AI로 자동 재탐색합니다.\n계속하시겠습니까?") \
                != QMessageBox.StandardButton.Yes:
            return
        self._on_save()
        self._refind_thread = BoardFinderThread(
            sites_path=self.sites_path, agency_filter=[agency],
            base_dir=self.base_dir, auto_save=True)
        self._refind_thread.result_ready.connect(
            lambda res: self._on_refind_done(agency, res))
        self._refind_thread.error.connect(
            lambda msg: QMessageBox.warning(self, "재탐색 오류", msg))
        self._refind_thread.finished.connect(
            lambda: (self.btn_refind.setText("AI 자동재탐색"),
                     self.btn_refind.setEnabled(True)))
        self.btn_refind.setEnabled(False)
        self.btn_refind.setText("탐색 중…")
        self._refind_thread.start()

    def _on_refind_done(self, agency, results):
        try:
            self._wb = openpyxl.load_workbook(self.sites_path)
            self._ws = self._wb.active
            self._agency_rows = {}
            for i, row in enumerate(list(self._ws.iter_rows(values_only=True))[1:], 2):
                if row and row[0]:
                    self._agency_rows.setdefault(str(row[0]).strip(), []).append(i)
        except Exception:
            pass
        self._rebuild_agency_list(keyword=self.agency_search.text())
        self._refresh_agency_color(agency)
        self._load_board_table(agency)
        self._dirty = False
        self._update_status()
        boards = results.get(agency, [])
        QMessageBox.information(
            self, "재탐색 완료", f"[{agency}] 게시판 {len(boards)}개 탐지됨.")

    # ─────────────── 엑셀 일괄가져오기
    def _on_bulk_import(self):
        if openpyxl is None:
            QMessageBox.warning(self, "오류", "openpyxl 패키지가 설치되어 있지 않습니다.")
            return
        try:
            path, _ = QFileDialog.getOpenFileName(
                self, "엑셀 가져오기", self.base_dir,
                "Excel 파일 (*.xlsx *.xls)")
            if not path:
                return
            wb_in = openpyxl.load_workbook(path, read_only=True)
            ws_in = wb_in.active
        except Exception as e:
            QMessageBox.warning(self, "가져오기 실패", f"파일 열기 오류: {e}")
            return

        rows = list(ws_in.iter_rows(values_only=True))
        if not rows:
            QMessageBox.warning(self, "가져오기", "파일이 비어있습니다.")
            return

        header = [str(c).strip() if c else "" for c in rows[0]]
        def ci(names):
            for n in names:
                for i, h in enumerate(header):
                    if n in h:
                        return i
            return None

        col_agency = ci(["기관명", "기관", "agency"])
        col_name   = ci(["게시판명", "게시판", "board"])
        col_url    = ci(["URL", "url", "주소"])
        col_param  = ci(["파라미터", "param", "페이지"])

        if None in (col_agency, col_name, col_url):
            QMessageBox.warning(self, "가져오기 실패",
                                "필수 컬럼(기관명/게시판명/URL)을 찾을 수 없습니다.\n"
                                "템플릿 다운로드 후 형식을 맞춰 주세요.")
            return

        added = skipped = 0
        unknown = []
        for data_row in rows[1:]:
            if not data_row or not data_row[0]:
                continue
            agency = str(data_row[col_agency] or "").strip()
            name   = str(data_row[col_name]   or "").strip()
            url    = str(data_row[col_url]    or "").strip()
            param  = str(data_row[col_param]  or "").strip() if col_param is not None else ""
            if not agency or not name or not url:
                skipped += 1
                continue
            if agency not in self._agency_rows:
                unknown.append(agency)
                skipped += 1
                continue
            self._write_board(agency, name, url, param, note="엑셀일괄")
            added += 1

        wb_in.close()
        self._dirty = True
        self._rebuild_agency_list(keyword=self.agency_search.text())
        self._update_status()

        msg = f"가져오기 완료: {added}개 추가, {skipped}개 건너뜀"
        if unknown:
            msg += f"\n미등록 기관(건너뜀): {', '.join(unknown[:5])}"
        QMessageBox.information(self, "가져오기 완료", msg)

    # ─────────────── 템플릿 다운로드
    def _on_download_template(self):
        if openpyxl is None:
            QMessageBox.warning(self, "오류", "openpyxl 패키지가 설치되어 있지 않습니다.")
            return
        try:
            save_path, _ = QFileDialog.getSaveFileName(
                self, "템플릿 저장",
                os.path.join(self.base_dir, "게시판_일괄등록_템플릿.xlsx"),
                "Excel 파일 (*.xlsx)")
            if not save_path:
                return
            wb_t = openpyxl.Workbook()
            ws_t = wb_t.active
            ws_t.title = "게시판목록"
            headers = ["기관명", "게시판명", "URL", "페이지파라미터"]
            navy = openpyxl.styles.PatternFill("solid", fgColor="1A3C6E")
            white_bold = openpyxl.styles.Font(bold=True, color="FFFFFF")
            for c, h in enumerate(headers, 1):
                cell = ws_t.cell(1, c, h)
                cell.font = white_bold
                cell.fill = navy
                ws_t.column_dimensions[
                    openpyxl.utils.get_column_letter(c)].width = 30

            for i, agency in enumerate(list(self._agency_rows.keys())[:5], 2):
                ws_t.cell(i, 1, agency)
                ws_t.cell(i, 2, "입찰공고")
                ws_t.cell(i, 3, "https://예시.go.kr/notice/list.do")
                ws_t.cell(i, 4, "?page=")

            ws_g = wb_t.create_sheet("작성안내")
            for r, row_data in enumerate([
                    ("컬럼명", "설명"),
                    ("기관명", "sites.xlsx에 등록된 기관명과 정확히 일치해야 합니다."),
                    ("게시판명", "게시판 또는 메뉴 이름. 예: 입찰공고"),
                    ("URL", "게시판 목록 첫 페이지 주소. https://로 시작"),
                    ("페이지파라미터", "페이지 이동 파라미터. 예: ?page=  없으면 빈칸"),
                ], 1):
                for c, val in enumerate(row_data, 1):
                    ws_g.cell(r, c, val)

            wb_t.save(save_path)
            QMessageBox.information(
                self, "템플릿 저장",
                f"저장 완료:\n{save_path}\n\n"
                "기관명·게시판명·URL 열을 채워 '엑셀로 일괄가져오기'를 사용하세요.")
        except Exception as e:
            QMessageBox.warning(self, "저장 실패", str(e))

    # ─────────────── 공통 쓰기 헬퍼
    def _write_board(self, agency, name, url, param="", note=""):
        rows = self._agency_rows.get(agency, [])
        target_row = None
        for r in rows:
            if not self._ws.cell(r, _COL["게시판명"]+1).value:
                target_row = r
                break
        if target_row is None:
            last = rows[-1] if rows else self._ws.max_row
            self._ws.insert_rows(last + 1)
            target_row = last + 1
            if rows:
                for col_key in ("기관명", "사이트타입", "활성화"):
                    self._ws.cell(target_row, _COL[col_key]+1).value = \
                        self._ws.cell(rows[0], _COL[col_key]+1).value
            self._agency_rows.setdefault(agency, []).append(target_row)
        self._ws.cell(target_row, _COL["게시판명"]+1).value    = name
        self._ws.cell(target_row, _COL["게시판URL"]+1).value   = url
        self._ws.cell(target_row, _COL["페이지파라미터"]+1).value = param
        self._ws.cell(target_row, _COL["비고"]+1).value        = note

    # ─────────────── 저장
    def _on_save(self):
        if not self._dirty or self._wb is None:
            return
        try:
            tmp = self.sites_path + ".tmp"
            self._wb.save(tmp)
            os.replace(tmp, self.sites_path)
            self._dirty = False
            self._update_status(saved=True)
        except PermissionError:
            QMessageBox.warning(
                self, "저장 실패",
                "sites.xlsx가 다른 프로그램에서 열려 있습니다.\n"
                "파일을 닫은 뒤 다시 저장하세요.")
        except Exception as e:
            QMessageBox.warning(self, "저장 실패", str(e))

    def _refresh_agency_color(self, agency):
        rows = self._agency_rows.get(agency, [])
        has_board = any(self._ws.cell(r, _COL["게시판명"]+1).value for r in rows)
        all_have  = all(self._ws.cell(r, _COL["게시판명"]+1).value for r in rows)
        for i in range(self.agency_list.count()):
            item = self.agency_list.item(i)
            if item.data(Qt.ItemDataRole.UserRole) == agency:
                if has_board and all_have:
                    item.setBackground(QBrush(COLOR_OK))
                elif has_board:
                    item.setBackground(QBrush(COLOR_PART))
                else:
                    item.setBackground(QBrush(COLOR_FAIL))
                cnt = sum(1 for r in rows if self._ws.cell(r, _COL["게시판명"]+1).value)
                item.setText(f"{agency}  ({cnt})")
                break

    def _update_status(self, saved=False):
        dirty_mark = ""
        if self._dirty:
            dirty_mark = "  [저장 필요]"
        elif saved:
            dirty_mark = f"  저장 완료 {datetime.now().strftime('%H:%M:%S')}"
        total = self.agency_list.count()
        ok   = sum(1 for i in range(total)
                   if self.agency_list.item(i).background().color() == COLOR_OK)
        fail = sum(1 for i in range(total)
                   if self.agency_list.item(i).background().color() == COLOR_FAIL)
        self.status_label.setText(
            f"전체 {total}  등록 {ok}  미등록 {fail}{dirty_mark}")

    def closeEvent(self, event):
        if self._dirty:
            ans = QMessageBox.question(
                self, "저장",
                "저장하지 않은 변경사항이 있습니다. 저장 후 닫으시겠습니까?",
                QMessageBox.StandardButton.Yes |
                QMessageBox.StandardButton.No  |
                QMessageBox.StandardButton.Cancel)
            if ans == QMessageBox.StandardButton.Cancel:
                event.ignore()
                return
            if ans == QMessageBox.StandardButton.Yes:
                self._on_save()
        super().closeEvent(event)
