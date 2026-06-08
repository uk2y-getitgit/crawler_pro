# -*- coding: utf-8 -*-
"""
board_finder.py — 안전점검 모니터링 시스템 Phase 4

66개 사이트의 메인 페이지에서 게시판 링크 목록을 수집하고,
BoardFinderAgent(AI)가 '공고/고시/입찰' 관련 게시판을 선별하여
sites.xlsx의 게시판명/게시판URL 컬럼을 채운다.

최초 1회 또는 신규 사이트 추가 시 실행한다.
requests + BeautifulSoup 기반 (JS 렌더링 사이트는 Phase 5의 playwright에서 처리).
"""

from __future__ import annotations

import abc
import json
import logging
import os
import sys
import time
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise ImportError("openpyxl가 필요합니다: pip install openpyxl") from e

# .env 로드 (선택)
try:
    from dotenv import load_dotenv
    _dotenv_available = True
except ImportError:
    _dotenv_available = False

# 선택적 SDK
try:
    import anthropic as _anthropic_sdk
    _anthropic_available = True
except ImportError:
    _anthropic_available = False

try:
    from google import genai as _genai
    from google.genai import types as _genai_types
    _gemini_available = True
except ImportError:
    _gemini_available = False

# loguru가 있으면 사용, 없으면 표준 logging
try:
    from loguru import logger  # type: ignore
except ImportError:
    logger = logging.getLogger("board_finder")
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        _h = logging.StreamHandler()
        _h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        logger.addHandler(_h)


# --------------------------------------------------------------------------- #
# BASE_DIR (EXE 호환) — crawler.py / ai_agent.py와 동일 패턴
# --------------------------------------------------------------------------- #
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # modules/board_finder.py 기준 -> project/
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

if _dotenv_available:
    _env_path = os.path.join(BASE_DIR, ".env")
    if os.path.exists(_env_path):
        load_dotenv(_env_path)


# --------------------------------------------------------------------------- #
# 상수
# --------------------------------------------------------------------------- #
# 초기 1회성 작업이므로 정확도 최우선 — haiku가 아닌 sonnet 사용
CLAUDE_MODEL = "claude-sonnet-4-20250514"
GEMINI_MODEL = "gemini-2.5-flash"

TIMEOUT = 10          # 접속 타임아웃 (초)
RETRY = 1             # 재시도 1회 (총 2회 시도)
FAIL_THRESHOLD = 2    # 연속 실패 임계치 -> 수동확인필요 표시

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

# AI 폴백 / 링크 사전필터에 쓰는 키워드
BOARD_INCLUDE_KEYWORDS = [
    "공고", "고시", "입찰", "공고문", "고시공고", "입찰공고",
    "지정공고", "조달", "구매", "계약", "발주", "낙찰", "용역",
]
BOARD_EXCLUDE_KEYWORDS = [
    "민원", "상담", "건의", "자유게시판", "자유",
    "소개", "기관안내", "조직도", "청사안내", "오시는", "찾아오시는",
    "보도자료", "보도", "뉴스", "갤러리", "사진", "동영상", "포토",
    "채용", "인사", "복지", "직원", "로그인", "회원", "검색",
    "사이트맵", "개인정보", "이용약관", "저작권", "rss",
]

# JS 렌더링이 의심되는 사이트 — 정적 추출 링크가 거의 없음.
# (정적 추출 결과가 매우 적으면 자동으로 'playwright필요'로 표시한다)
MIN_LINKS_FOR_STATIC = 3


SYSTEM_PROMPT = """당신은 대한민국 관공서 웹사이트 게시판 분류 전문가입니다.
아래 메뉴 링크 목록에서 '공고/고시/입찰' 관련 게시판만 선별하세요.

【반드시 포함】
- 공고, 고시, 입찰공고, 공고문, 고시공고, 지정공고
- 조달/구매/계약 관련 공고

【반드시 제외】
- 민원, 상담, 건의, 자유게시판
- 소개, 기관안내, 조직도, 청사안내
- 보도자료, 뉴스, 갤러리, 사진, 동영상
- 채용, 인사, 복지, 직원

【출력 형식】JSON만 출력, 설명 없음:
{"selected": [{"name": "게시판명", "url": "https://..."}], "reason": "선택 이유 한 줄"}"""


# --------------------------------------------------------------------------- #
# 내부 예외
# --------------------------------------------------------------------------- #
class _APIFatalError(Exception):
    """API 치명 오류 — 키워드 폴백으로 전환해야 함."""


class _ParseError(Exception):
    """AI 응답 JSON 파싱 실패 — 키워드 폴백으로 전환."""


# --------------------------------------------------------------------------- #
# 유틸 (ai_agent.py와 동일한 키 검증 로직)
# --------------------------------------------------------------------------- #
def _is_valid_claude_key(key: str) -> bool:
    if not key:
        return False
    key = key.strip()
    if not key.startswith("sk-ant-"):
        return False
    markers = ("여기에", "입력", "your", "xxx", "placeholder", "...")
    low = key.lower()
    if any(m in key or m in low for m in markers):
        return False
    return len(key) >= 30


def _is_valid_gemini_key(key: str) -> bool:
    if not key:
        return False
    key = key.strip()
    markers = ("여기에", "입력", "your", "xxx", "placeholder", "...")
    low = key.lower()
    if any(m in key or m in low for m in markers):
        return False
    return len(key) >= 20


def _parse_json_response(text: str) -> dict:
    """AI 응답 텍스트에서 JSON 객체를 추출한다. 실패 시 _ParseError."""
    if not text:
        raise _ParseError("빈 응답")
    cleaned = text.strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    # ```json ... ``` 코드펜스 / 앞뒤 설명 제거
    try:
        c = cleaned
        if c.startswith("```"):
            parts = c.split("```")
            c = parts[1] if len(parts) >= 2 else cleaned
            if c.lower().startswith("json"):
                c = c[4:]
            c = c.strip()
        start, end = c.find("{"), c.rfind("}")
        if start != -1 and end > start:
            return json.loads(c[start:end + 1])
    except (json.JSONDecodeError, IndexError):
        pass
    raise _ParseError("JSON 파싱 실패")


def _keyword_fallback_select(links: list[dict]) -> dict:
    """
    텍스트 기반 키워드 폴백.
    - exclude 키워드 포함 → 제외
    - include 키워드 포함 → 선별
    """
    selected = []
    seen = set()
    for link in links:
        text = (link.get("text") or "").strip()
        url = (link.get("url") or "").strip()
        if not text or not url:
            continue
        low = text.lower()
        if any(e and (e in text or e in low) for e in BOARD_EXCLUDE_KEYWORDS):
            continue
        if any(k in text for k in BOARD_INCLUDE_KEYWORDS):
            if url in seen:
                continue
            seen.add(url)
            selected.append({"name": text, "url": url})
    return {
        "selected": selected,
        "reason": f"키워드 폴백 선별 ({len(selected)}건)",
    }


def _build_user_message(links: list[dict]) -> str:
    lines = ["다음 메뉴 링크 목록에서 공고/고시/입찰 게시판을 선별해주세요:"]
    for i, link in enumerate(links):
        text = (link.get("text") or "").strip()
        url = (link.get("url") or "").strip()
        lines.append(f"[{i}] 메뉴명: {text} | URL: {url}")
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# BoardFinderAgent (AI)
# --------------------------------------------------------------------------- #
class _BaseBoardAgent(abc.ABC):
    provider_name = "base"

    def __init__(self):
        self.use_fallback = False
        self.fallback_reason = ""

    @abc.abstractmethod
    def _call_api(self, user_msg: str) -> str:
        """API 호출 — 응답 텍스트 반환. 치명 오류 시 _APIFatalError."""

    def find_boards(self, links: list[dict]) -> dict:
        """
        links: [{"text": "메뉴명", "url": "https://..."}]
        반환: {"selected": [{"name": ..., "url": ...}], "reason": "..."}
        AI 실패/파싱 실패 시 키워드 폴백으로 자동 전환.
        """
        if not links:
            return {"selected": [], "reason": "링크 없음"}
        if self.use_fallback:
            return _keyword_fallback_select(links)
        try:
            text = self._call_api(_build_user_message(links))
            data = _parse_json_response(text)
            return self._normalize(data, links)
        except (_APIFatalError, _ParseError) as e:
            logger.warning(f"[{self.provider_name}] AI 실패, 키워드 폴백 전환: {e}")
            self.use_fallback = True
            self.fallback_reason = str(e)
            return _keyword_fallback_select(links)
        except Exception as e:  # 예기치 못한 오류도 폴백으로
            logger.warning(f"[{self.provider_name}] 예외, 키워드 폴백 전환: {e}")
            return _keyword_fallback_select(links)

    @staticmethod
    def _normalize(data: dict, links: list[dict]) -> dict:
        """AI 응답을 정규화. selected의 각 항목이 name/url을 갖도록 보정."""
        raw = data.get("selected", []) if isinstance(data, dict) else []
        valid_urls = {(l.get("url") or "").strip() for l in links}
        selected = []
        seen = set()
        for item in raw:
            if not isinstance(item, dict):
                continue
            name = (item.get("name") or "").strip()
            url = (item.get("url") or "").strip()
            if not url:
                continue
            # AI가 환각으로 만든 URL은 원본 링크에 있는 것만 신뢰
            if valid_urls and url not in valid_urls:
                continue
            if url in seen:
                continue
            seen.add(url)
            selected.append({"name": name or url, "url": url})
        reason = ""
        if isinstance(data, dict):
            reason = (data.get("reason") or "").strip()
        return {"selected": selected, "reason": reason or "AI 선별"}


class _ClaudeBoardAgent(_BaseBoardAgent):
    provider_name = "claude"

    def __init__(self, api_key: str | None = None):
        super().__init__()
        self.api_key = api_key or os.getenv("ANTHROPIC_API_KEY", "")
        self.client = None
        if not _anthropic_available:
            self.use_fallback = True
            self.fallback_reason = "anthropic SDK 미설치"
        elif not _is_valid_claude_key(self.api_key):
            self.use_fallback = True
            self.fallback_reason = "ANTHROPIC_API_KEY 없음 또는 미설정"
        else:
            try:
                self.client = _anthropic_sdk.Anthropic(api_key=self.api_key)
            except Exception as e:
                self.use_fallback = True
                self.fallback_reason = f"클라이언트 초기화 실패: {e}"
        if self.use_fallback:
            logger.warning(f"[ClaudeBoardAgent] {self.fallback_reason}")

    def _call_api(self, user_msg: str) -> str:
        try:
            resp = self.client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=2048,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": user_msg}],
            )
            for block in resp.content:
                if getattr(block, "type", None) == "text":
                    return block.text
            return ""
        except (_anthropic_sdk.AuthenticationError,
                _anthropic_sdk.PermissionDeniedError) as e:
            raise _APIFatalError(f"인증 실패: {e}") from e
        except _anthropic_sdk.APIError as e:
            raise _APIFatalError(f"API 오류: {e}") from e


class _GeminiBoardAgent(_BaseBoardAgent):
    provider_name = "gemini"

    def __init__(self, api_key: str | None = None):
        super().__init__()
        self.api_key = api_key or os.getenv("GEMINI_API_KEY", "")
        self.client = None
        if not _gemini_available:
            self.use_fallback = True
            self.fallback_reason = "google-genai SDK 미설치"
        elif not _is_valid_gemini_key(self.api_key):
            self.use_fallback = True
            self.fallback_reason = "GEMINI_API_KEY 없음 또는 미설정"
        else:
            try:
                self.client = _genai.Client(api_key=self.api_key)
            except Exception as e:
                self.use_fallback = True
                self.fallback_reason = f"클라이언트 초기화 실패: {e}"
        if self.use_fallback:
            logger.warning(f"[GeminiBoardAgent] {self.fallback_reason}")

    def _call_api(self, user_msg: str) -> str:
        try:
            resp = self.client.models.generate_content(
                model=GEMINI_MODEL,
                contents=user_msg,
                config=_genai_types.GenerateContentConfig(
                    system_instruction=SYSTEM_PROMPT,
                    max_output_tokens=2048,
                ),
            )
            return resp.text or ""
        except Exception as e:
            err = str(e).lower()
            if any(k in err for k in ("api_key_invalid", "permission",
                                      "unauthorized", "403")):
                raise _APIFatalError(f"인증 실패: {e}") from e
            raise _APIFatalError(f"Gemini API 오류: {e}") from e


def create_board_finder_agent(provider: str | None = None) -> _BaseBoardAgent:
    """
    AI provider에 맞는 BoardFinderAgent를 생성한다.
    create_post_filter_agent()와 동일한 전환 로직:
      provider=None → .env의 AI_PROVIDER (기본 claude)
      지정 provider 실패 → 반대 provider 자동 시도 → 키워드 폴백.
    """
    if provider is None:
        provider = os.getenv("AI_PROVIDER", "claude").lower().strip()

    if provider == "gemini":
        agent = _GeminiBoardAgent()
        if not agent.use_fallback:
            logger.info("[AI] GeminiBoardAgent 활성화")
            return agent
        logger.warning(f"[AI] Gemini 실패 ({agent.fallback_reason}), Claude 시도")
        alt = _ClaudeBoardAgent()
        if not alt.use_fallback:
            logger.info("[AI] ClaudeBoardAgent 로 자동 전환")
            return alt
    else:
        agent = _ClaudeBoardAgent()
        if not agent.use_fallback:
            logger.info("[AI] ClaudeBoardAgent 활성화")
            return agent
        logger.warning(f"[AI] Claude 실패 ({agent.fallback_reason}), Gemini 시도")
        alt = _GeminiBoardAgent()
        if not alt.use_fallback:
            logger.info("[AI] GeminiBoardAgent 로 자동 전환")
            return alt

    logger.warning("[AI] Claude·Gemini 모두 초기화 실패, 키워드 폴백 모드")
    agent.use_fallback = True
    agent.fallback_reason = "Claude·Gemini API 키 모두 미설정 — 키워드 폴백 사용"
    return agent


# 하위 호환 별칭 (명세상 BoardFinderAgent 이름 노출)
class BoardFinderAgent(_ClaudeBoardAgent):
    """하위 호환 별칭. 새 코드는 create_board_finder_agent() 사용 권장."""
    pass


# --------------------------------------------------------------------------- #
# 링크 추출 헬퍼
# --------------------------------------------------------------------------- #
def _same_domain(base_url: str, target_url: str) -> bool:
    """target_url이 base_url과 같은 도메인(또는 서브도메인)인지 확인."""
    try:
        b = urlparse(base_url).netloc.lower()
        t = urlparse(target_url).netloc.lower()
    except ValueError:
        return False
    if not t:
        return False  # netloc 없음 → 비정상
    if b == t:
        return True
    # 서브도메인 허용 (예: www.x.go.kr ↔ open.x.go.kr 의 공통 등록 도메인)
    b_root = ".".join(b.split(".")[-3:]) if b.count(".") >= 2 else b
    t_root = ".".join(t.split(".")[-3:]) if t.count(".") >= 2 else t
    return b_root == t_root


def _normalize_link(base_url: str, href: str) -> str | None:
    """
    href를 절대 URL로 정규화한다.
    제외 대상이면 None 반환:
      - 빈 값 / 앵커만(#) / javascript: / mailto: / tel:
      - 외부 도메인
    """
    if not href:
        return None
    href = href.strip()
    if not href:
        return None
    low = href.lower()
    # 앵커만
    if href == "#" or href.startswith("#"):
        return None
    # javascript / mailto / tel 등 비-HTTP 스킴
    if low.startswith(("javascript:", "mailto:", "tel:", "data:")):
        return None
    # 절대 URL로 변환
    full = urljoin(base_url, href)
    parsed = urlparse(full)
    if parsed.scheme not in ("http", "https"):
        return None
    # 외부 링크 제외
    if not _same_domain(base_url, full):
        return None
    return full


# --------------------------------------------------------------------------- #
# BoardFinder (메인 클래스)
# --------------------------------------------------------------------------- #
class BoardFinder:
    # sites.xlsx 컬럼 인덱스 (1-base, crawler._load_sites와 동일 순서)
    COL_AGENCY = 1       # 기관명
    COL_SITE_TYPE = 2    # 사이트타입
    COL_BOARD_NAME = 3   # 게시판명
    COL_BOARD_URL = 4    # 게시판URL
    COL_PAGE_PARAM = 5   # 페이지파라미터
    COL_ACTIVE = 6       # 활성화
    COL_NOTE = 7         # 비고

    def __init__(self, sites_path, progress_callback=None, base_dir=None):
        self.base_dir = base_dir or BASE_DIR
        self.sites_path = self._abspath(sites_path)
        self.progress_callback = progress_callback or (lambda et, data: None)
        self.agent = create_board_finder_agent()
        self._fail_counts = {}

    # ------------------------------------------------------------------ utils
    def _abspath(self, path):
        if os.path.isabs(path):
            return path
        return os.path.join(self.base_dir, path)

    def _emit(self, event_type, data):
        try:
            self.progress_callback(event_type, data)
        except Exception as e:
            logger.warning(f"progress_callback 오류: {e}")

    # ------------------------------------------------------------------ sites
    def _load_sites(self):
        wb = openpyxl.load_workbook(self.sites_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0] if rows else ()
        sites = []
        for idx, row in enumerate(rows[1:], start=2):  # 엑셀 행 번호(2부터)
            if row is None or all(c is None for c in row):
                continue
            sites.append({
                "row": idx,
                "agency": row[0],
                "site_type": row[1],
                "board_name": row[2],
                "url": row[3],
                "page_param": row[4],
                "active": row[5],
                "note": row[6] if len(row) > 6 else None,
            })
        return wb, ws, header, sites

    # ---------------------------------------------------------------- fetch
    def _fetch(self, url):
        """타임아웃 10초, 재시도 1회, SSL 오류 시 verify=False 재시도."""
        headers = {
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
        }
        last_err = None
        for attempt in range(RETRY + 1):
            try:
                resp = requests.get(url, headers=headers, timeout=TIMEOUT,
                                    verify=True)
                return resp.status_code, resp.text, None
            except requests.exceptions.SSLError:
                try:
                    resp = requests.get(url, headers=headers, timeout=TIMEOUT,
                                        verify=False)
                    return resp.status_code, resp.text, None
                except Exception as e2:
                    last_err = f"SSL/{e2}"
            except requests.exceptions.RequestException as e:
                last_err = str(e)
            if attempt < RETRY:
                time.sleep(1)
        return None, None, last_err

    def fetch_links(self, url: str) -> list[dict]:
        """
        메인 페이지 접속 → a태그 링크+텍스트 추출.
        반환: [{"text": "메뉴명", "url": "절대URL"}]
        - JS 렌더링 의심 사이트(정적 링크 빈약)는 빈 리스트 반환
          (호출부에서 비고 'playwright필요' 표시)
        """
        if not url:
            return []
        status, html, err = self._fetch(url)
        if err is not None:
            logger.warning(f"  접속 실패: {err}")
            return []
        if status is None or status >= 400:
            logger.warning(f"  HTTP {status}")
            return []
        if not html:
            return []

        soup = BeautifulSoup(html, "lxml")
        links = []
        seen = set()
        for a in soup.find_all("a"):
            href = a.get("href", "")
            text = a.get_text(strip=True)
            full = _normalize_link(url, href)
            if full is None:
                continue
            if not text or len(text) < 2:
                continue
            key = (text, full)
            if key in seen:
                continue
            seen.add(key)
            links.append({"text": text, "url": full})
        return links

    # ----------------------------------------------------------------- run
    def run_single(self, agency: str) -> list[dict]:
        """
        단일 기관 처리: fetch_links → BoardFinderAgent.find_boards.
        반환: [{"name": ..., "url": ...}]  (선별된 게시판 목록)
        """
        wb, ws, header, sites = self._load_sites()
        target = next((s for s in sites if s["agency"] == agency), None)
        if target is None:
            logger.warning(f"기관 '{agency}'을(를) sites.xlsx에서 찾을 수 없음")
            return []
        return self._process_site(target)

    def _process_site(self, site: dict) -> list[dict]:
        agency = site["agency"]
        url = site["url"]
        self._emit("start_site", {"agency": agency, "url": url})
        logger.info(f"{agency} 게시판 탐색: {url}")

        if not url:
            self._emit("error", {"agency": agency, "error": "URL 없음"})
            return []

        links = self.fetch_links(url)
        if len(links) < MIN_LINKS_FOR_STATIC:
            # 정적 추출 실패 / JS 렌더링 의심
            cnt = self._fail_counts.get(agency, 0) + 1
            self._fail_counts[agency] = cnt
            logger.warning(f"  {agency}: 추출 링크 {len(links)}개 (정적 추출 빈약)")
            self._emit("error", {
                "agency": agency,
                "error": "링크없음" if not links else "playwright필요",
                "link_count": len(links),
            })
            return []

        self._fail_counts[agency] = 0
        result = self.agent.find_boards(links)
        selected = result.get("selected", [])
        self._emit("site_done", {
            "agency": agency,
            "link_count": len(links),
            "selected_count": len(selected),
            "reason": result.get("reason", ""),
        })
        logger.info(f"  {agency}: {len(links)}개 링크 중 {len(selected)}개 게시판 선별")
        return selected

    def run_all(self, agency_filter=None) -> dict:
        """
        전체 또는 지정 기관 처리.
        agency_filter=None → sites.xlsx 전체
        agency_filter=문자열 또는 리스트 → 해당 기관만
        반환: {기관명: [{"name": ..., "url": ...}]}
        """
        if isinstance(agency_filter, str):
            agency_filter = [agency_filter]

        wb, ws, header, sites = self._load_sites()
        target = []
        for s in sites:
            if agency_filter is not None and s["agency"] not in agency_filter:
                continue
            target.append(s)

        results = {}
        for i, site in enumerate(target):
            self._emit("progress", {"index": i + 1, "total": len(target)})
            selected = self._process_site(site)
            results[site["agency"]] = selected
        self._emit("complete", {
            "total_agencies": len(target),
            "with_boards": sum(1 for v in results.values() if v),
        })
        return results

    # ------------------------------------------------------------ save sheet
    def _append_note(self, cell, tag):
        """비고 셀에 tag를 중복 없이 추가."""
        existing = cell.value or ""
        if tag in str(existing):
            return
        cell.value = (str(existing) + " " + tag).strip() if existing else tag

    def save_to_sites(self, results: dict):
        """
        결과를 sites.xlsx의 게시판명/게시판URL 컬럼에 저장한다.

        주의: 이 프로젝트에서 게시판URL 컬럼은 board_finder가 탐색에 사용하는
        '시드 메인 URL'로 모든 행에 미리 채워져 있다. 따라서 '기존 값 보존'의
        기준은 게시판URL 유무가 아니라 '게시판명(컬럼3)이 이미 채워졌는지'로 판단한다.
        (게시판명이 비어 있으면 아직 board_finder가 정리하지 않은 시드 상태)

        - 게시판명이 이미 채워진 행 → 덮어쓰지 않고 비고에 '확인필요' 표시
        - 1개 기관에 여러 게시판 → 첫 게시판은 기존 행에, 나머지는 행 분리하여 추가
        - 선별 결과 0개 → 비고에 '링크없음' 표시
        """
        wb, ws, header, sites = self._load_sites()
        # 기관명 → 원본 행 매핑 (첫 등장 행 사용)
        agency_row = {}
        for s in sites:
            agency_row.setdefault(s["agency"], s)

        added_rows = 0
        for agency, boards in results.items():
            base_site = agency_row.get(agency)
            if base_site is None:
                logger.warning(f"저장 스킵: '{agency}' 행 없음")
                continue
            row_idx = base_site["row"]
            note_cell = ws.cell(row=row_idx, column=self.COL_NOTE)

            if not boards:
                self._append_note(note_cell, "링크없음")
                continue

            existing_name = ws.cell(row=row_idx, column=self.COL_BOARD_NAME).value
            if existing_name and str(existing_name).strip():
                # 이미 board_finder가 정리한 행 — 덮어쓰지 않고 확인 필요 표시
                self._append_note(note_cell, "확인필요")
                logger.info(f"  {agency}: 게시판명 기존 값 존재 → '확인필요' 표시")
                continue

            # 첫 게시판은 기존 행에 기록
            first = boards[0]
            ws.cell(row=row_idx, column=self.COL_BOARD_NAME).value = first.get("name", "")
            ws.cell(row=row_idx, column=self.COL_BOARD_URL).value = first.get("url", "")

            # 나머지 게시판은 새 행으로 분리 추가
            for extra in boards[1:]:
                new_row = ws.max_row + 1
                ws.cell(row=new_row, column=self.COL_AGENCY).value = agency
                ws.cell(row=new_row, column=self.COL_SITE_TYPE).value = base_site.get("site_type") or "일반"
                ws.cell(row=new_row, column=self.COL_BOARD_NAME).value = extra.get("name", "")
                ws.cell(row=new_row, column=self.COL_BOARD_URL).value = extra.get("url", "")
                ws.cell(row=new_row, column=self.COL_PAGE_PARAM).value = base_site.get("page_param")
                ws.cell(row=new_row, column=self.COL_ACTIVE).value = base_site.get("active") or "Y"
                added_rows += 1

        try:
            wb.save(self.sites_path)
            logger.info(f"sites.xlsx 저장 완료 (행 {added_rows}개 추가)")
        except PermissionError:
            logger.error("sites.xlsx 저장 실패: 파일이 열려 있습니다. 닫고 다시 시도하세요.")
            raise
        return added_rows

    # ---------------------------------------------------------- interactive
    def interactive_review(self, results: dict) -> dict:
        """
        CLI에서 AI 선별 결과를 보여주고 사람이 확인/수정/스킵.
        각 기관별로:
          [Enter] 확정 / s 스킵(저장 제외) / e 수정(번호 입력으로 일부 선택)
        반환: 최종 확정된 결과 {기관명: [...]}
        """
        final = {}
        for agency, boards in results.items():
            print("\n" + "=" * 60)
            print(f"[{agency}]  AI 선별 게시판 {len(boards)}개")
            if not boards:
                print("  (선별된 게시판 없음)")
            for i, b in enumerate(boards):
                print(f"  {i+1}. {b.get('name', '')}  ->  {b.get('url', '')}")
            print("-" * 60)
            print("  [Enter] 확정  |  s: 스킵  |  e: 번호로 선택(예: 1,3)")
            try:
                choice = input("  선택> ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                print("\n중단됨 — 지금까지 확정된 결과만 반환합니다.")
                break

            if choice == "s":
                print(f"  {agency} 스킵")
                continue
            if choice == "e":
                try:
                    raw = input("  포함할 번호(콤마 구분)> ").strip()
                except (EOFError, KeyboardInterrupt):
                    raw = ""
                picked = []
                for tok in raw.replace(" ", "").split(","):
                    if tok.isdigit():
                        n = int(tok) - 1
                        if 0 <= n < len(boards):
                            picked.append(boards[n])
                final[agency] = picked
                print(f"  {agency} {len(picked)}개 확정")
            else:
                final[agency] = boards
                print(f"  {agency} {len(boards)}개 확정")
        return final
