# -*- coding: utf-8 -*-
"""
crawler.py — 안전점검 모니터링 시스템 Phase 2

sites.xlsx에서 활성화=Y인 사이트의 게시글 목록을 수집한다.
requests + BeautifulSoup 기반 (playwright는 Phase 5).
"""

import os
import sys
import json
import time
import logging
from datetime import datetime, timedelta
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

# verify=False 우회 시 발생하는 InsecureRequestWarning 소음 억제
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception:
    pass

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise ImportError("openpyxl가 필요합니다: pip install openpyxl") from e

# .env 로드 (선택)
try:
    from dotenv import load_dotenv
    _dotenv_available = True
except ImportError:
    _dotenv_available = False

# loguru가 있으면 사용, 없으면 표준 logging
try:
    from loguru import logger  # type: ignore
    _USE_LOGURU = True
except ImportError:
    _USE_LOGURU = False
    logger = logging.getLogger("crawler")
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        _h = logging.StreamHandler()
        _h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        logger.addHandler(_h)


# ---------------------------------------------------------------------------
# BASE_DIR (EXE 호환)
# ---------------------------------------------------------------------------
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # modules/crawler.py 기준 -> project/
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# .env 로드
if _dotenv_available:
    _env_path = os.path.join(BASE_DIR, ".env")
    if os.path.exists(_env_path):
        load_dotenv(_env_path)


def _env_int(name, default):
    try:
        return int(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        return default


CRAWL_DELAY = _env_int("CRAWL_DELAY", 2)
MAX_PAGES = _env_int("MAX_PAGES", 3)
SEARCH_DAYS = _env_int("SEARCH_DAYS", 3)  # 게시일이 최근 N일 이내인 공고만 통과

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

TIMEOUT = 15          # 초
RETRY = 2             # 재시도 2회 (총 3회 시도, 각 시도마다 verify True/False)
FAIL_THRESHOLD = 3    # 연속 실패 임계치 -> 점검필요 표시
HISTORY_MONTHS = 6    # 6개월 지난 항목 삭제


# CSS 선택자 자동 탐지 순서
SELECTOR_STRATEGY = [
    "table.bbs_list tr td a",
    "table tr td a",
    "ul.board_list li a",
    "div.board_list a",
    '[class*="list"] a[href]',
]


# 제목 뒤에 붙은 메타정보(작성자/작성일 등) 절단용 — 행 전체를 긁은 경우 정제
_TITLE_CUT_MARKERS = ("작성자", "작성일", "등록일", "게시일", "조회수", "조회 ",
                      "첨부", "담당자", "담당부서", "작성부서")
# 제목 앞 라벨 제거 대상
_TITLE_LEAD_LABELS = ("제목", "제목:", "제목 :", "공지", "[공지]")


def _clean_title(raw_text, anchor=None):
    """게시글 제목을 정제한다.
    - anchor의 title 속성이 더 길면 그것을 사용(목록 잘림 보완)
    - 선두 '제목/공지' 라벨 제거
    - '작성자/작성일' 등 메타정보가 뒤에 붙은 경우(행 전체를 긁음) 절단
    """
    t = (raw_text or "").strip()
    if anchor is not None:
        try:
            attr = (anchor.get("title") or "").strip()
        except Exception:
            attr = ""
        if len(attr) > len(t):
            t = attr
    # 연속 공백 정리
    t = " ".join(t.split())
    # 선두 라벨 제거
    for lead in _TITLE_LEAD_LABELS:
        if t.startswith(lead):
            t = t[len(lead):].strip(" :\t-")
            break
    # 메타정보 절단 (제목이 어느 정도 있은 뒤 등장할 때만)
    for m in _TITLE_CUT_MARKERS:
        idx = t.find(m)
        if idx > 4:
            t = t[:idx].strip()
    return t.strip()


def _is_valid_post_link(text, href):
    """제목 링크로 보이는지 휴리스틱 검사.

    href가 javascript:fnView(...) 형태여도 제목이 유효하면 게시글로 인정한다.
    (관공서 게시판 다수가 onclick/javascript로 상세를 여는데, 이를 버리면
    게시글이 0건이 된다. URL은 _extract_posts에서 게시판 목록 URL로 대체한다.)
    """
    text = (text or "").strip()
    if len(text) < 3:
        return False
    # 페이지네이션/메뉴성 텍스트 제외
    lowered = text.lower()
    skip_words = ("다음", "이전", "처음", "마지막", "more", "목록", "검색", "로그인",
                  "home", "더보기", "바로가기", "메뉴", "닫기", "전체보기", "rss")
    if lowered in skip_words or text in ("1", "2", "3", "4", "5"):
        return False
    if text.replace(",", "").isdigit():   # 번호/조회수 등 순수 숫자 셀 제외
        return False
    if any(s in text for s in ("바로가기",)):
        return False
    return True


def _is_js_href(href):
    """href가 직접 이동 불가한 형태(javascript/앵커/빈값)인지."""
    h = (href or "").strip()
    return (not h) or h == "#" or h.startswith("#") or \
        h.lower().startswith("javascript")


class Crawler:
    def __init__(self, sites_path, history_path, progress_callback=None,
                 base_dir=None, ai_agent=None):
        self.base_dir = base_dir or BASE_DIR
        self.sites_path = self._abspath(sites_path)
        self.history_path = self._abspath(history_path)
        self.progress_callback = progress_callback or (lambda et, data: None)
        # AI 필터 (PostFilterAgent). None이면 키워드 폴백을 사용한다.
        self.ai_agent = ai_agent

        self.errors = []          # 오류로그용
        self.excluded = []        # 제외 게시글(날짜/AI) + 사유 — 추적·감사용
        self.history = self._load_history()
        # 연속 실패 카운트 추적용 (메모리 내)
        self._fail_counts = {}
        # 심화(Playwright) 페처 — 첫 심화 사이트에서 지연 생성, run 종료 시 정리
        self._deep_fetcher = None
        # 중단 플래그 (GUI/스레드에서 stop() 호출 시 set). 사이트 경계에서만
        # 검사하므로 진행 중인 단일 사이트 수집·파일 저장이 중간에 끊기지 않는다.
        self._stop_requested = False

    def request_stop(self):
        """크롤링 중단을 요청한다. 현재 사이트 처리를 마친 뒤 안전하게 종료된다."""
        self._stop_requested = True

    # ------------------------------------------------------------------ utils
    def _abspath(self, path):
        if os.path.isabs(path):
            return path
        return os.path.join(self.base_dir, path)

    def _emit(self, event_type, data):
        try:
            self.progress_callback(event_type, data)
        except Exception as e:  # 콜백 오류가 크롤링을 멈추면 안 됨
            logger.warning(f"progress_callback 오류: {e}")

    # ---------------------------------------------------------------- history
    def _load_history(self):
        try:
            with open(self.history_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                raise ValueError("history.json 형식 오류")
            data.setdefault("last_updated", "")
            data.setdefault("items", {})
            return data
        except (FileNotFoundError, json.JSONDecodeError, ValueError) as e:
            logger.warning(f"history.json 로드 실패, 새로 생성: {e}")
            return {"last_updated": "", "items": {}}

    def _prune_history(self):
        """6개월 지난 항목 삭제."""
        cutoff = datetime.now() - timedelta(days=HISTORY_MONTHS * 30)
        items = self.history.get("items", {})
        removed = 0
        for key in list(items.keys()):
            date_str = items[key].get("date", "")
            dt = _parse_date(date_str)
            if dt is not None and dt < cutoff:
                del items[key]
                removed += 1
        if removed:
            logger.info(f"history 정리: {removed}개 항목 삭제 (6개월 경과)")

    def _save_history(self):
        self.history["last_updated"] = datetime.now().isoformat()
        os.makedirs(os.path.dirname(self.history_path), exist_ok=True)
        with open(self.history_path, "w", encoding="utf-8") as f:
            json.dump(self.history, f, ensure_ascii=False, indent=2)

    @staticmethod
    def _make_key(title, date, url):
        if url:
            return url
        return f"{title}_{date}"

    @staticmethod
    def _within_date_window(date_str):
        """게시일이 최근 SEARCH_DAYS일 이내인지 판정.
        반환 (통과여부, 사유). 날짜를 못 읽으면 통과시킨다(누락 방지)."""
        if SEARCH_DAYS <= 0:
            return True, ""
        dt = _parse_date(date_str)
        if dt is None:
            return True, "날짜미상(포함)"
        cutoff = datetime.now() - timedelta(days=SEARCH_DAYS)
        # 날짜 단위 비교 (시각 무시)
        if dt.date() >= cutoff.date():
            return True, ""
        return False, f"게시일 {dt.strftime('%Y-%m-%d')} ({SEARCH_DAYS}일 초과)"

    def _record_excluded(self, base, reason):
        """제외 게시글을 사유와 함께 기록한다."""
        item = dict(base)
        item["exclude_reason"] = reason
        item["ai_reason"] = reason
        self.excluded.append(item)

    # ------------------------------------------------------------------ sites
    def _load_sites(self):
        wb = openpyxl.load_workbook(self.sites_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        sites = []
        for idx, row in enumerate(rows[1:], start=2):  # 엑셀 행 번호(2부터)
            if row is None or all(c is None for c in row):
                continue
            site = {
                "row": idx,
                "agency": row[0],
                "site_type": row[1],
                "board_name": row[2],
                "url": row[3],
                "page_param": row[4],
                "active": row[5],
                "note": row[6],
            }
            sites.append(site)
        return wb, ws, header, sites

    # ---------------------------------------------------------------- fetch
    def _fetch(self, url):
        """타임아웃 10초, 재시도 1회. (status_code, html|None, error|None)"""
        headers = {
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
        }
        last_err = None
        # 매 시도마다 verify=True → 실패 시 verify=False로도 시도한다.
        # 정부서버는 GPKI 인증서 문제(SSLError)뿐 아니라 핸드셰이크 중
        # 연결 리셋(ConnectionError)도 잦은데, verify=False로 우회되는 경우가 많다.
        for attempt in range(RETRY + 1):
            for verify in (True, False):
                try:
                    resp = requests.get(url, headers=headers, timeout=TIMEOUT,
                                        verify=verify)
                    return resp.status_code, resp.text, None
                except requests.exceptions.RequestException as e:
                    last_err = str(e)
                    continue
            if attempt < RETRY:
                time.sleep(1.5 * (attempt + 1))  # 지수 백오프
        return None, None, last_err

    @staticmethod
    def _is_deep(site):
        """site_type이 '심화'면 Playwright 경로를 사용한다."""
        return str(site.get("site_type") or "").strip() == "심화"

    def _get_deep_fetcher(self):
        """심화 페처를 지연 생성하여 재사용한다(브라우저 1회 기동)."""
        if self._deep_fetcher is None:
            from modules.playwright_crawler import DeepFetcher  # 지연 임포트
            df = DeepFetcher(headless=True)
            df.start()
            self._deep_fetcher = df
            logger.info("  심화 엔진(Playwright/Chromium) 기동")
        return self._deep_fetcher

    def _close_deep_fetcher(self):
        if self._deep_fetcher is not None:
            try:
                self._deep_fetcher.close()
            except Exception as e:
                logger.warning(f"심화 페처 종료 오류: {e}")
            self._deep_fetcher = None

    def _fetch_deep(self, url):
        """심화 사이트를 브라우저로 렌더링. (status, html, error) — _fetch와 동일 시그니처."""
        try:
            df = self._get_deep_fetcher()
            status, html, final, err = df.fetch(url)
            if err is not None:
                return None, None, err
            return status, html, None
        except Exception as e:
            return None, None, f"심화엔진오류: {type(e).__name__}: {str(e)[:100]}"

    # ------------------------------------------------------------- extract
    def _extract_posts(self, html, base_url):
        """CSS 선택자 전략 순서대로 시도. (posts, used_selector)"""
        soup = BeautifulSoup(html, "lxml")
        for selector in SELECTOR_STRATEGY:
            try:
                anchors = soup.select(selector)
            except Exception:
                continue
            posts = []
            seen = set()
            for a in anchors:
                raw = a.get_text(strip=True)
                href = a.get("href", "")
                if not _is_valid_post_link(raw, href):
                    continue
                title = _clean_title(raw, a)
                if len(title) < 3:
                    continue
                # javascript/앵커 링크는 상세 직접이동이 불가 → 게시판 목록 URL로 대체.
                # (제목 기반 AI 판정은 그대로 동작하고, 사용자는 목록에서 원문 확인)
                if _is_js_href(href):
                    full_url = base_url
                    key = ("js", title)        # 목록URL 공유 → 제목으로 중복제거
                else:
                    full_url = urljoin(base_url, href)
                    key = full_url
                if key in seen:
                    continue
                seen.add(key)
                date = _find_date_near(a)
                posts.append({"title": title, "url": full_url, "date": date})
            if posts:
                return posts, selector
        return [], None

    def _board_pages(self, site):
        """1..MAX_PAGES 페이지 URL 생성."""
        base = site["url"]
        param = site.get("page_param")
        urls = [base]
        if param:
            for p in range(2, MAX_PAGES + 1):
                sep = "&" if "?" in base else "?"
                urls.append(f"{base}{sep}{param}={p}")
        return urls

    # ----------------------------------------------------------------- run
    def run(self, site_filter=None):
        """
        site_filter: None이면 활성화=Y 전체, 리스트면 해당 기관명만.
        반환: 결과 dict 리스트.
        """
        self._prune_history()
        wb, ws, header, sites = self._load_sites()

        target = []
        for s in sites:
            if s["active"] != "Y":
                continue
            if site_filter is not None and s["agency"] not in site_filter:
                continue
            target.append(s)

        results = []
        sheet_dirty = False

        stopped = False
        for i, site in enumerate(target):
            # 사이트 경계에서 중단 요청 확인 (진행 중 사이트는 끊지 않음)
            if self._stop_requested:
                stopped = True
                logger.info("중단 요청 감지 — 현재까지 수집한 결과로 마무리합니다.")
                self._emit("stopped", {"completed": i, "total": len(target)})
                break

            agency = site["agency"]
            board_name = site["board_name"] or ""
            url = site["url"]

            self._emit("start_site", {
                "index": i + 1, "total": len(target),
                "agency": agency, "url": url,
            })
            logger.info(f"[{i+1}/{len(target)}] {agency} 수집 시작: {url}")

            if not url:
                self._record_error(agency, url, "게시판URL 없음", 0, "스킵")
                self._emit("error", {"agency": agency, "error": "URL 없음"})
                continue

            site_posts, err = self._crawl_site(site)

            if err is not None:
                # 실패 처리 + 연속 실패 카운트
                cnt = self._fail_counts.get(agency, 0) + 1
                self._fail_counts[agency] = cnt
                self._record_error(agency, url, err, RETRY, "실패")
                self._emit("error", {"agency": agency, "error": err, "fail_count": cnt})
                logger.warning(f"  {agency} 실패: {err} (연속 {cnt}회)")
                if cnt >= FAIL_THRESHOLD:
                    self._mark_inspection_needed(ws, site)
                    sheet_dirty = True
            else:
                self._fail_counts[agency] = 0
                if not site_posts:
                    logger.warning(f"  {agency}: 게시글 0개 — 구조 변경 의심")
                    self._record_error(agency, url, "게시글 0개 (선택자 미매칭)",
                                       RETRY, "경고")
                for post in site_posts:
                    key = self._make_key(post["title"], post["date"], post["url"])
                    is_new = key not in self.history["items"]
                    base = {
                        "agency": agency,
                        "board_name": board_name,
                        "title": post["title"],
                        "date": post["date"],
                        "url": post["url"],
                        "is_new": is_new,
                        "ai_reason": "",
                    }

                    # 게시일 조건 검사 — 기간 밖이면 제외(사유 기록), 결과에서 빼고
                    # history에는 기록해 다음 실행 때 재처리하지 않는다.
                    within, date_reason = self._within_date_window(post["date"])
                    if not within:
                        if is_new:
                            self._record_excluded(base, date_reason)
                            self.history["items"][key] = {
                                "title": post["title"], "date": post["date"],
                                "agency": agency,
                            }
                        continue

                    results.append(base)
                    if is_new:
                        self.history["items"][key] = {
                            "title": post["title"],
                            "date": post["date"],
                            "agency": agency,
                        }
                        self._emit("post_found", base)

            # 사이트 간 딜레이 (마지막 사이트 제외).
            # 중단 요청이 들어오면 즉시 빠져나올 수 있도록 1초 단위로 끊어 잔다.
            if i < len(target) - 1:
                for _ in range(CRAWL_DELAY):
                    if self._stop_requested:
                        break
                    time.sleep(1)

        # 점검필요 표시가 있었으면 sites.xlsx 저장
        if sheet_dirty:
            try:
                wb.save(self.sites_path)
                logger.info("sites.xlsx 비고란 '점검필요' 저장 완료")
            except Exception as e:
                logger.warning(f"sites.xlsx 저장 실패: {e}")

        # 심화 브라우저 정리(수집 종료 후 더 이상 불필요)
        self._close_deep_fetcher()

        # AI 필터 적용: 수집한 게시글 중 '안전점검 수행기관 지정공고'만 선별
        collected = len(results)
        candidates = results
        results = self._apply_ai_filter(candidates)
        # AI가 제외한 게시글(통과하지 못한 것)을 사유와 함께 기록한다.
        matched_ids = {id(m) for m in results}
        for r in candidates:
            if id(r) not in matched_ids and r.get("is_new"):
                self._record_excluded(r, "AI 제외: 안전점검 수행기관 지정공고 아님")

        self._save_history()
        self._emit("complete", {
            "total_collected": collected,
            "matched_count": len(results),
            "new_count": sum(1 for r in results if r["is_new"]),
            "error_count": len(self.errors),
            "stopped": stopped,
        })
        logger.info(f"수집 완료: 총 {collected}건 수집, "
                    f"AI 선별 {len(results)}건, "
                    f"신규 {sum(1 for r in results if r['is_new'])}건, "
                    f"오류 {len(self.errors)}건")
        return results

    def _apply_ai_filter(self, results):
        """
        수집된 게시글에 AI 필터를 적용한다.
        - ai_agent가 있으면 PostFilterAgent.filter() 호출
        - 없으면 키워드 폴백 필터(keyword_fallback_filter) 사용
        선별된 게시글만 ai_reason이 채워진 채로 반환된다.
        """
        if not results:
            return results
        try:
            if self.ai_agent is not None:
                filtered = self.ai_agent.filter(results)
            else:
                # 지연 임포트: ai_agent 미사용 환경에서도 crawler 단독 동작.
                # 패키지/단독 실행 양쪽 import 방식 모두 지원.
                try:
                    from modules.ai_agent import keyword_fallback_filter
                except ImportError:
                    from ai_agent import keyword_fallback_filter
                filtered = keyword_fallback_filter(results)
            self._emit("ai_filter_done", {
                "before": len(results), "after": len(filtered),
            })
            return filtered
        except Exception as e:
            # 필터 자체가 실패하면 수집 결과를 그대로 반환(데이터 유실 방지)
            logger.warning(f"AI 필터 적용 실패, 원본 반환: {e}")
            self._emit("error", {"agency": "AI필터", "error": str(e)})
            return results

    def _crawl_site(self, site):
        """단일 사이트 전체 페이지 수집. (posts, error)"""
        deep = self._is_deep(site)
        all_posts = []
        seen_urls = set()
        for page_url in self._board_pages(site):
            if deep:
                status, html, err = self._fetch_deep(page_url)
            else:
                status, html, err = self._fetch(page_url)
            if err is not None:
                # 첫 페이지 실패면 사이트 실패, 2페이지 이후 실패는 무시
                if page_url == site["url"]:
                    return [], err
                else:
                    logger.warning(f"  {site['agency']} 페이지 실패(무시): {err}")
                    break
            if status in (404, 503) or (status is not None and status >= 400):
                if page_url == site["url"]:
                    return [], f"HTTP {status}"
                else:
                    logger.warning(f"  {site['agency']} HTTP {status} (페이지 스킵)")
                    break
            posts, selector = self._extract_posts(html, page_url)
            if selector:
                logger.info(f"  선택자 매칭: '{selector}' -> {len(posts)}건")
            for p in posts:
                # javascript 게시글은 URL이 목록주소로 동일 → 제목까지 묶어 중복제거
                dedup_key = (p["title"], p["url"])
                if dedup_key in seen_urls:
                    continue
                seen_urls.add(dedup_key)
                all_posts.append(p)
        return all_posts, None

    # ----------------------------------------------------- error / sheet ops
    def _record_error(self, agency, url, message, retries, outcome):
        self.errors.append({
            "agency": agency,
            "url": url or "",
            "error": message,
            "retries": retries,
            "outcome": outcome,
        })

    def _mark_inspection_needed(self, ws, site):
        """sites.xlsx 비고(7번째 컬럼, G열)에 '점검필요' 표시."""
        try:
            cell = ws.cell(row=site["row"], column=7)
            existing = cell.value or ""
            if "점검필요" not in str(existing):
                cell.value = (str(existing) + " 점검필요").strip() if existing \
                    else "점검필요"
            logger.warning(f"  {site['agency']} 연속 {FAIL_THRESHOLD}회 실패 "
                           f"-> 비고 '점검필요' 표시")
        except Exception as e:
            logger.warning(f"비고 표시 실패: {e}")


# ---------------------------------------------------------------------------
# 날짜 파싱 헬퍼
# ---------------------------------------------------------------------------
_DATE_FORMATS = [
    "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d",
    "%Y년 %m월 %d일", "%y-%m-%d", "%y.%m.%d",
    "%Y-%m-%d %H:%M", "%Y.%m.%d %H:%M",
]

import re

_DATE_RE = re.compile(
    r"(20\d{2})\s*[-./년]\s*(\d{1,2})\s*[-./월]\s*(\d{1,2})"
)


def _parse_date(date_str):
    if not date_str:
        return None
    date_str = date_str.strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    m = _DATE_RE.search(date_str)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _normalize_date(raw):
    """추출한 날짜 문자열을 YYYY-MM-DD로 정규화. 실패 시 원문 반환."""
    dt = _parse_date(raw)
    if dt:
        return dt.strftime("%Y-%m-%d")
    return raw.strip() if raw else ""


def _find_date_near(anchor):
    """앵커 주변(같은 tr/li, 형제 td 등)에서 날짜 문자열을 찾는다."""
    # 1) 부모 행(tr/li) 전체 텍스트에서 정규식 탐색
    container = anchor.find_parent(["tr", "li"])
    if container is None:
        container = anchor.parent
    if container is not None:
        text = container.get_text(" ", strip=True)
        m = _DATE_RE.search(text)
        if m:
            return _normalize_date(m.group(0))
    # 2) 앵커 자체 텍스트
    m = _DATE_RE.search(anchor.get_text(" ", strip=True))
    if m:
        return _normalize_date(m.group(0))
    return ""
