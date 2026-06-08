# -*- coding: utf-8 -*-
"""
gui_main.py — 안전점검 모니터링 시스템
메인 윈도우. CrawlerThread / SchedulerThread / 다이얼로그를 연결한다.
"""

from __future__ import annotations

import os
import logging

from PyQt6.QtCore import Qt, QSortFilterProxyModel
from PyQt6.QtGui import QAction, QColor, QDesktopServices, QFont
from PyQt6.QtCore import QUrl
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QToolBar, QTreeWidget, QTreeWidgetItem,
    QTableWidget, QTableWidgetItem, QPlainTextEdit, QSplitter, QVBoxLayout,
    QHBoxLayout, QPushButton, QLabel, QFileDialog, QMessageBox, QHeaderView,
    QAbstractItemView, QLineEdit, QComboBox, QFrame,
)

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from modules.crawler_thread import CrawlerThread
    from modules.gui_dialogs import (SettingsDialog, BoardFinderDialog,
                                     BoardManageDialog, _read_env)
    from modules.scheduler import SchedulerThread
    from modules.reporter import Reporter
except ImportError:
    from crawler_thread import CrawlerThread
    from gui_dialogs import (SettingsDialog, BoardFinderDialog,
                              BoardManageDialog, _read_env)
    from scheduler import SchedulerThread
    from reporter import Reporter

try:
    from loguru import logger
except ImportError:
    logger = logging.getLogger("gui_main")


SITES_REL   = os.path.join("config", "sites.xlsx")
HISTORY_REL = os.path.join("data", "history.json")

COLOR_NEW    = QColor("#FFF9C4")   # 신규 — 연노랑
COLOR_HEADER = QColor("#1a3c6e")

# ─────────────────────────────────────────────────────────
# 스타일시트
# ─────────────────────────────────────────────────────────
STYLESHEET = """
/* ── 전체 기반 ── */
QMainWindow, QDialog, QWidget {
    background-color: #F0F4F8;
    color: #2C3E50;
    font-family: '맑은 고딕', 'Apple SD Gothic Neo', 'Segoe UI', sans-serif;
    font-size: 13px;
}

/* ── 툴바 ── */
QToolBar {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #1e4d8c, stop:1 #163a6e);
    spacing: 3px;
    padding: 5px 10px;
    border: none;
}
QToolBar QToolButton {
    color: #FFFFFF;
    background-color: transparent;
    padding: 7px 16px;
    font-size: 13px;
    font-weight: 600;
    border-radius: 5px;
    border: 1.5px solid transparent;
    min-width: 68px;
}
QToolBar QToolButton:hover {
    background-color: rgba(255,255,255,0.18);
    border-color: rgba(255,255,255,0.45);
}
QToolBar QToolButton:pressed {
    background-color: rgba(0,0,0,0.22);
}
QToolBar QToolButton:disabled {
    color: rgba(255,255,255,0.32);
}
QToolBar::separator {
    width: 1px;
    background: rgba(255,255,255,0.25);
    margin: 6px 4px;
}

/* ── 카드 패널 ── */
QFrame#card {
    background: #FFFFFF;
    border: 1px solid #DDE3EA;
    border-radius: 6px;
}

/* ── 필터 바 ── */
QFrame#filterBar {
    background: #FFFFFF;
    border: 1px solid #DDE3EA;
    border-radius: 6px;
}

/* ── 트리·테이블·로그 ── */
QTreeWidget, QTableWidget, QPlainTextEdit {
    background-color: #FFFFFF;
    border: 1px solid #DDE3EA;
    color: #2C3E50;
    border-radius: 4px;
    selection-background-color: #D6E4F7;
    selection-color: #1a3c6e;
    alternate-background-color: #F7F9FC;
}
QTreeWidget::item, QTableWidget::item { padding: 3px 4px; }
QTreeWidget::item:hover, QTableWidget::item:hover {
    background-color: #EBF3FD;
}
QHeaderView::section {
    background-color: #1a3c6e;
    color: #FFFFFF;
    padding: 6px 8px;
    border: none;
    font-weight: bold;
    font-size: 12px;
}

/* ── 버튼 ── */
QPushButton {
    background-color: #1a3c6e;
    color: #FFFFFF;
    padding: 6px 15px;
    border-radius: 5px;
    font-weight: 600;
    border: none;
    min-height: 28px;
}
QPushButton:hover  { background-color: #265fa3; }
QPushButton:pressed { background-color: #122d56; }
QPushButton:disabled { background-color: #B0BEC5; color: #ECEFF1; }

QPushButton[secondary="true"] {
    background-color: #FFFFFF;
    color: #1a3c6e;
    border: 1.5px solid #1a3c6e;
}
QPushButton[secondary="true"]:hover { background-color: #EBF3FD; }

QPushButton[danger="true"] {
    background-color: #E53935;
}
QPushButton[danger="true"]:hover { background-color: #C62828; }

QPushButton[success="true"] {
    background-color: #2E7D32;
}
QPushButton[success="true"]:hover { background-color: #1B5E20; }

/* ── 입력·콤보 ── */
QLineEdit, QComboBox {
    border: 1.5px solid #CBD5E1;
    border-radius: 4px;
    padding: 5px 9px;
    background: #FFFFFF;
    color: #2C3E50;
    min-height: 26px;
}
QLineEdit:focus, QComboBox:focus { border-color: #1a3c6e; }
QComboBox::drop-down { border: none; width: 22px; }

/* ── 상태바 ── */
QStatusBar {
    background-color: #E8EEF7;
    color: #1a3c6e;
    border-top: 1px solid #DDE3EA;
    padding: 2px 8px;
}

/* ── 그룹박스 ── */
QGroupBox {
    border: 1.5px solid #DDE3EA;
    border-radius: 6px;
    margin-top: 10px;
    padding-top: 10px;
    font-weight: bold;
    color: #1a3c6e;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 6px;
}

/* ── 라벨 ── */
QLabel { color: #2C3E50; }
QLabel[title="true"] { font-size: 14px; font-weight: bold; color: #1a3c6e; }
QLabel[chip="true"] {
    background: #EBF3FD;
    color: #1a3c6e;
    border-radius: 10px;
    padding: 2px 10px;
    font-size: 12px;
    font-weight: bold;
}
QLabel[chip_new="true"] {
    background: #FFF9C4;
    color: #7B6200;
    border-radius: 10px;
    padding: 2px 10px;
    font-size: 12px;
    font-weight: bold;
}
"""

RESULT_HEADERS = ["신규", "기관명", "게시판명", "게시글 제목", "게시일", "URL"]
MAX_LOG_LINES  = 600


class MainWindow(QMainWindow):
    def __init__(self, base_dir=None):
        super().__init__()
        self.base_dir    = base_dir or os.path.dirname(
            os.path.dirname(os.path.abspath(__file__)))
        self.sites_path   = os.path.join(self.base_dir, SITES_REL)
        self.history_path = os.path.join(self.base_dir, HISTORY_REL)

        self._crawler_thread   = None
        self._scheduler_thread = None
        self._results  = []
        self._errors   = []
        self._excluded = []   # 제외 게시글(날짜/AI) — 엑셀 제외공고 시트용

        self.setWindowTitle("안전점검 수행기관 모니터링")
        self.resize(1180, 760)
        self.setStyleSheet(STYLESHEET)

        self._build_toolbar()
        self._build_central()
        self._build_statusbar()
        self._load_sites_tree()
        self._maybe_start_scheduler()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 툴바
    def _build_toolbar(self):
        tb = QToolBar("메인")
        tb.setMovable(False)
        tb.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
        self.addToolBar(tb)

        # 수집 그룹
        self.act_run = QAction("전체 수집", self)
        self.act_run.setToolTip("등록된 전체 활성 사이트에서 공고를 수집합니다")
        self.act_run.triggered.connect(self._run_all)
        tb.addAction(self.act_run)

        self.act_run_sel = QAction("선택 수집", self)
        self.act_run_sel.setToolTip("체크한 사이트만 수집합니다")
        self.act_run_sel.triggered.connect(self._run_selected)
        tb.addAction(self.act_run_sel)

        self.act_stop = QAction("수집 중지", self)
        self.act_stop.setToolTip("진행 중인 수집을 안전하게 멈춥니다")
        self.act_stop.triggered.connect(self._stop_crawl)
        self.act_stop.setEnabled(False)
        tb.addAction(self.act_stop)

        tb.addSeparator()

        # 게시판 관리 그룹
        self.act_board_find = QAction("게시판 자동탐색", self)
        self.act_board_find.setToolTip("AI로 등록된 사이트에서 게시판을 자동 탐색합니다")
        self.act_board_find.triggered.connect(self._open_board_finder)
        tb.addAction(self.act_board_find)

        self.act_board_manage = QAction("게시판 관리", self)
        self.act_board_manage.setToolTip("게시판 목록 조회 / 수동 추가·수정·삭제")
        self.act_board_manage.triggered.connect(self._open_board_manage)
        tb.addAction(self.act_board_manage)

        tb.addSeparator()

        self.act_settings = QAction("환경설정", self)
        self.act_settings.setToolTip("API 키, 크롤 딜레이, 자동 스케줄 설정")
        self.act_settings.triggered.connect(self._open_settings)
        tb.addAction(self.act_settings)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 중앙 레이아웃
    def _build_central(self):
        # ── 사이트 목록 (왼쪽) ──
        left_frame = QFrame()
        left_frame.setObjectName("card")
        left_frame.setFixedWidth(240)
        ll = QVBoxLayout(left_frame)
        ll.setContentsMargins(8, 8, 8, 8)
        ll.setSpacing(6)

        tree_title = QLabel("사이트 목록")
        tree_title.setProperty("title", True)
        ll.addWidget(tree_title)

        self.tree = QTreeWidget()
        self.tree.setHeaderHidden(True)
        self.tree.setAlternatingRowColors(True)
        ll.addWidget(self.tree)

        chk_row = QHBoxLayout()
        btn_check_all = QPushButton("전체 선택")
        btn_check_all.setProperty("secondary", True)
        btn_check_all.clicked.connect(lambda: self._set_all_check(True))
        btn_uncheck_all = QPushButton("전체 해제")
        btn_uncheck_all.setProperty("secondary", True)
        btn_uncheck_all.clicked.connect(lambda: self._set_all_check(False))
        chk_row.addWidget(btn_check_all)
        chk_row.addWidget(btn_uncheck_all)
        ll.addLayout(chk_row)

        # ── 오른쪽 (필터 + 테이블) ──
        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(6)

        # 필터 바
        rl.addWidget(self._build_filter_bar())

        # 결과 테이블
        result_frame = QFrame()
        result_frame.setObjectName("card")
        rf_layout = QVBoxLayout(result_frame)
        rf_layout.setContentsMargins(8, 8, 8, 8)
        rf_layout.setSpacing(4)

        result_header = QHBoxLayout()
        lbl = QLabel("수집 결과")
        lbl.setProperty("title", True)
        result_header.addWidget(lbl)
        result_header.addStretch()
        self.lbl_total = QLabel("0건")
        self.lbl_total.setProperty("chip", True)
        self.lbl_new   = QLabel("신규 0건")
        self.lbl_new.setProperty("chip_new", True)
        result_header.addWidget(self.lbl_total)
        result_header.addWidget(self.lbl_new)
        btn_load = QPushButton("최근 결과 불러오기")
        btn_load.setProperty("secondary", True)
        btn_load.setFixedWidth(140)
        btn_load.setToolTip("results 폴더의 가장 최근 결과 엑셀을 표에 불러옵니다")
        btn_load.clicked.connect(self._load_latest_results)
        result_header.addWidget(btn_load)
        btn_excel = QPushButton("엑셀 저장")
        btn_excel.setProperty("success", True)
        btn_excel.setFixedWidth(90)
        btn_excel.clicked.connect(self._save_excel)
        result_header.addWidget(btn_excel)
        rf_layout.addLayout(result_header)

        self.table = QTableWidget(0, len(RESULT_HEADERS))
        self.table.setHorizontalHeaderLabels(RESULT_HEADERS)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(5, QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(5, 200)
        self.table.cellDoubleClicked.connect(self._open_url)
        rf_layout.addWidget(self.table)
        rl.addWidget(result_frame)

        # ── 수평 분할 ──
        top_split = QSplitter(Qt.Orientation.Horizontal)
        top_split.addWidget(left_frame)
        top_split.addWidget(right)
        top_split.setStretchFactor(0, 0)
        top_split.setStretchFactor(1, 1)

        # ── 로그 영역 ──
        log_frame = QFrame()
        log_frame.setObjectName("card")
        lf_layout = QVBoxLayout(log_frame)
        lf_layout.setContentsMargins(8, 6, 8, 6)
        lf_layout.setSpacing(4)

        log_header = QHBoxLayout()
        log_header.addWidget(QLabel("실행 로그"))
        log_header.addStretch()
        btn_clear_log = QPushButton("로그 지우기")
        btn_clear_log.setProperty("secondary", True)
        btn_clear_log.setFixedHeight(24)
        btn_clear_log.clicked.connect(lambda: self.log.clear())
        log_header.addWidget(btn_clear_log)
        lf_layout.addLayout(log_header)

        self.log = QPlainTextEdit()
        self.log.setReadOnly(True)
        self.log.setMaximumBlockCount(MAX_LOG_LINES)
        lf_layout.addWidget(self.log)
        log_frame.setMinimumHeight(130)
        log_frame.setMaximumHeight(200)

        # ── 수직 분할 ──
        main_split = QSplitter(Qt.Orientation.Vertical)
        main_split.addWidget(top_split)
        main_split.addWidget(log_frame)
        main_split.setStretchFactor(0, 1)
        main_split.setStretchFactor(1, 0)
        main_split.setSizes([560, 160])

        container = QWidget()
        cl = QVBoxLayout(container)
        cl.setContentsMargins(8, 6, 8, 6)
        cl.setSpacing(6)
        cl.addWidget(main_split)
        self.setCentralWidget(container)

    def _build_filter_bar(self):
        """결과 필터 바를 생성한다."""
        bar = QFrame()
        bar.setObjectName("filterBar")
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(10, 6, 10, 6)
        layout.setSpacing(8)

        # 기관명 검색
        layout.addWidget(QLabel("기관명"))
        self.filter_agency = QLineEdit()
        self.filter_agency.setPlaceholderText("기관명 검색…")
        self.filter_agency.setFixedWidth(150)
        self.filter_agency.textChanged.connect(self._apply_filters)
        layout.addWidget(self.filter_agency)

        # 제목 검색
        layout.addWidget(QLabel("제목"))
        self.filter_title = QLineEdit()
        self.filter_title.setPlaceholderText("게시글 제목 검색…")
        self.filter_title.setFixedWidth(190)
        self.filter_title.textChanged.connect(self._apply_filters)
        layout.addWidget(self.filter_title)

        # 게시판 필터
        layout.addWidget(QLabel("게시판"))
        self.filter_board = QComboBox()
        self.filter_board.addItem("전체")
        self.filter_board.setFixedWidth(155)
        self.filter_board.currentIndexChanged.connect(self._apply_filters)
        layout.addWidget(self.filter_board)

        # 신규만 토글
        self.btn_new_only = QPushButton("신규만")
        self.btn_new_only.setCheckable(True)
        self.btn_new_only.setFixedWidth(70)
        self.btn_new_only.setProperty("secondary", True)
        self.btn_new_only.toggled.connect(self._apply_filters)
        layout.addWidget(self.btn_new_only)

        # 필터 초기화
        btn_reset = QPushButton("초기화")
        btn_reset.setProperty("secondary", True)
        btn_reset.setFixedWidth(65)
        btn_reset.clicked.connect(self._clear_filters)
        layout.addWidget(btn_reset)

        layout.addStretch()

        self.lbl_filter_count = QLabel("결과 0건")
        self.lbl_filter_count.setStyleSheet("color:#1a3c6e; font-weight:bold;")
        layout.addWidget(self.lbl_filter_count)

        return bar

    def _build_statusbar(self):
        self.status = self.statusBar()
        self.status.showMessage("대기 중")

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 트리
    def _load_sites_tree(self):
        self.tree.clear()
        if openpyxl is None or not os.path.exists(self.sites_path):
            self._log("sites.xlsx를 찾을 수 없습니다.")
            return
        try:
            wb = openpyxl.load_workbook(self.sites_path, read_only=True)
            ws = wb.active
            seen_agencies: dict[str, QTreeWidgetItem] = {}
            board_count:  dict[str, int] = {}

            for row in list(ws.iter_rows(values_only=True))[1:]:
                if not row or row[0] is None:
                    continue
                agency  = str(row[0]).strip()
                bname   = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                active  = str(row[5]).strip().upper() == "Y" if len(row) > 5 and row[5] else False

                if agency not in seen_agencies:
                    item = QTreeWidgetItem()
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    item.setCheckState(
                        0, Qt.CheckState.Checked if active else Qt.CheckState.Unchecked)
                    seen_agencies[agency] = item
                    board_count[agency]   = 0
                    self.tree.addTopLevelItem(item)

                if bname:
                    board_count[agency] = board_count.get(agency, 0) + 1

            # 이름 업데이트 (게시판 수 포함)
            for agency, item in seen_agencies.items():
                cnt = board_count.get(agency, 0)
                item.setText(0, f"{agency}  ({cnt}개)" if cnt else f"⚠ {agency}")
                if cnt == 0:
                    item.setForeground(0, QColor("#C0392B"))
            wb.close()
        except Exception as e:
            self._log(f"사이트 목록 로드 실패: {e}")

    def _checked_agencies(self):
        out = []
        for i in range(self.tree.topLevelItemCount()):
            it = self.tree.topLevelItem(i)
            if it.checkState(0) == Qt.CheckState.Checked:
                # 이름에서 기관명만 추출 (앞부분)
                name = it.text(0).split("  (")[0].lstrip("⚠ ")
                out.append(name)
        return out

    def _set_all_check(self, state: bool):
        cs = Qt.CheckState.Checked if state else Qt.CheckState.Unchecked
        for i in range(self.tree.topLevelItemCount()):
            self.tree.topLevelItem(i).setCheckState(0, cs)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 수집
    def _run_all(self):
        self._start_crawl(site_filter=None)

    def _run_selected(self):
        agencies = self._checked_agencies()
        if not agencies:
            QMessageBox.information(self, "선택 수집", "체크된 사이트가 없습니다.")
            return
        self._start_crawl(site_filter=agencies)

    def _start_crawl(self, site_filter):
        if self._crawler_thread is not None and self._crawler_thread.isRunning():
            QMessageBox.information(self, "수집 중", "이미 수집이 진행 중입니다.")
            return
        self._results = []
        self._errors  = []
        self.table.setRowCount(0)
        self._update_counts()
        self._set_running(True)
        scope = "전체 활성 사이트" if site_filter is None else f"{len(site_filter)}개 사이트"
        self._log(f"━ 수집 시작: {scope} ━")
        self.status.showMessage("수집 준비 중…")

        self._crawler_thread = CrawlerThread(
            sites_path=self.sites_path,
            history_path=self.history_path,
            site_filter=site_filter,
            base_dir=self.base_dir,
        )
        self._crawler_thread.progress.connect(self._on_progress)
        self._crawler_thread.result_ready.connect(self._on_crawl_done)
        self._crawler_thread.error.connect(self._on_crawl_error)
        self._crawler_thread.finished.connect(self._on_thread_finished)
        self._crawler_thread.start()

    def _stop_crawl(self):
        if self._crawler_thread is not None and self._crawler_thread.isRunning():
            self._log("중지 요청 — 현재 처리 완료 후 멈춥니다.")
            self.status.showMessage("중지 중…")
            self._crawler_thread.stop()
            self.act_stop.setEnabled(False)

    def _set_running(self, running: bool):
        self.act_run.setEnabled(not running)
        self.act_run_sel.setEnabled(not running)
        self.act_board_find.setEnabled(not running)
        self.act_board_manage.setEnabled(not running)
        self.act_stop.setEnabled(running)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 스레드 시그널
    def _on_progress(self, event_type, data):
        if event_type == "start_site":
            idx    = data.get("index", "?")
            total  = data.get("total", "?")
            agency = data.get("agency", "")
            self.status.showMessage(f"수집 중 ({idx}/{total}) — {agency}")
            self._log(f"[{idx}/{total}] {agency}")
        elif event_type == "post_found":
            self._results.append(data)
            self._apply_filters()
        elif event_type == "error":
            self._log(f"  ⚠ 오류: {data.get('agency','')}: {data.get('error','')}")
        elif event_type == "ai_filter_done":
            self._log(f"  AI 선별: {data.get('before')}건 → {data.get('after')}건")
        elif event_type == "ai_fallback":
            self._log(f"  AI 폴백: {data.get('reason','')}")
        elif event_type == "stopped":
            self._log(f"  중단 ({data.get('completed')}/{data.get('total')} 완료)")

    def _on_crawl_done(self, results, errors):
        self._results = results
        self._errors  = errors
        self._excluded = getattr(self._crawler_thread, "excluded", []) or []
        self._update_board_filter_combo()
        self._apply_filters()
        new_count = sum(1 for r in results if r.get("is_new"))
        ex_count = len(self._excluded)
        msg = (f"완료 — 통과 {len(results)}건 (신규 {new_count}) / "
               f"제외 {ex_count} / 오류 {len(errors)}")
        self.status.showMessage(msg)
        self._log(f"━ {msg} ━")

    def _on_crawl_error(self, msg):
        self._log(f"[치명 오류] {msg}")
        self.status.showMessage("오류로 중단됨")
        QMessageBox.critical(self, "수집 오류", msg)

    def _on_thread_finished(self):
        self._set_running(False)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 테이블 / 필터
    def _apply_filters(self, *_):
        agency_kw = self.filter_agency.text().strip().lower()
        title_kw  = self.filter_title.text().strip().lower()
        board_sel = self.filter_board.currentText()
        new_only  = self.btn_new_only.isChecked()

        filtered = self._results
        if agency_kw:
            filtered = [r for r in filtered
                        if agency_kw in r.get("agency", "").lower()]
        if title_kw:
            filtered = [r for r in filtered
                        if title_kw in r.get("title", "").lower()]
        if board_sel and board_sel != "전체":
            filtered = [r for r in filtered
                        if r.get("board_name", "") == board_sel]
        if new_only:
            filtered = [r for r in filtered if r.get("is_new")]

        self.table.setRowCount(0)
        for r in filtered:
            self._render_row(self.table.rowCount(), r)

        self.lbl_filter_count.setText(f"결과 {len(filtered)}건")
        self._update_counts()

    def _clear_filters(self):
        self.filter_agency.clear()
        self.filter_title.clear()
        self.filter_board.setCurrentIndex(0)
        self.btn_new_only.setChecked(False)

    def _update_board_filter_combo(self):
        boards = sorted(set(
            r.get("board_name", "") for r in self._results if r.get("board_name")))
        current = self.filter_board.currentText()
        self.filter_board.blockSignals(True)
        self.filter_board.clear()
        self.filter_board.addItem("전체")
        for b in boards:
            if b:
                self.filter_board.addItem(b)
        idx = self.filter_board.findText(current)
        self.filter_board.setCurrentIndex(idx if idx >= 0 else 0)
        self.filter_board.blockSignals(False)

    def _update_counts(self):
        total    = len(self._results)
        new_cnt  = sum(1 for r in self._results if r.get("is_new"))
        self.lbl_total.setText(f"총 {total}건")
        self.lbl_new.setText(f"신규 {new_cnt}건")

    def _render_row(self, row: int, item: dict):
        self.table.insertRow(row)
        is_new = bool(item.get("is_new"))
        values = [
            "★ 신규" if is_new else "기존",
            item.get("agency", ""),
            item.get("board_name", ""),
            item.get("title", ""),
            item.get("date", ""),
            item.get("url", ""),
        ]
        for c, val in enumerate(values):
            cell = QTableWidgetItem(str(val))
            if is_new:
                cell.setBackground(COLOR_NEW)
            if c == 0 and is_new:
                cell.setForeground(QColor("#7B6200"))
                f = cell.font()
                f.setBold(True)
                cell.setFont(f)
            if c == 5 and str(val).startswith("http"):
                # URL 컬럼 — 링크처럼 표시(파란색·밑줄) + 안내 툴팁
                cell.setForeground(QColor("#0563C1"))
                f = cell.font()
                f.setUnderline(True)
                cell.setFont(f)
                cell.setToolTip("더블클릭하면 원문 게시글을 브라우저로 엽니다")
            self.table.setItem(row, c, cell)

    def _open_url(self, row, col):
        url_item = self.table.item(row, 5)
        if url_item and url_item.text().startswith("http"):
            QDesktopServices.openUrl(QUrl(url_item.text()))

    def _load_latest_results(self):
        """results 폴더의 가장 최근 결과 엑셀(신규공고 시트)을 표에 불러온다.
        재크롤 없이 통과 게시글 + 원문 URL을 바로 확인할 수 있다."""
        import glob
        results_dir = os.path.join(self.base_dir, "results")
        files = glob.glob(os.path.join(results_dir, "결과_*.xlsx"))
        if not files:
            QMessageBox.information(self, "최근 결과", "results 폴더에 결과 파일이 없습니다.")
            return
        latest = max(files, key=os.path.getmtime)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
            ws = wb["신규공고"] if "신규공고" in wb.sheetnames else wb.active
            loaded = []
            for row in list(ws.iter_rows(values_only=True))[1:]:
                if not row or not any(row):
                    continue
                # POST_HEADERS: 신규여부,수집일시,기관명,게시판명,게시글제목,게시일,URL,AI판단근거
                loaded.append({
                    "is_new": str(row[0]).strip() == "신규",
                    "agency": row[2] or "", "board_name": row[3] or "",
                    "title": row[4] or "", "date": row[5] or "",
                    "url": row[6] or "", "ai_reason": row[7] or "",
                })
            wb.close()
        except Exception as e:
            QMessageBox.warning(self, "불러오기 실패", f"결과 파일 읽기 오류:\n{e}")
            return
        self._results = loaded
        self._excluded = []
        self._update_board_filter_combo()
        self._apply_filters()
        self._log(f"최근 결과 불러옴: {os.path.basename(latest)} ({len(loaded)}건)")
        self.status.showMessage(
            f"최근 결과 {len(loaded)}건 불러옴 — {os.path.basename(latest)}")

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 엑셀 저장
    def _save_excel(self):
        if not self._results:
            QMessageBox.information(self, "엑셀 저장", "저장할 결과가 없습니다.")
            return
        default = os.path.join(self.base_dir, "results")
        path, _ = QFileDialog.getSaveFileName(
            self, "엑셀 저장", os.path.join(default, "결과.xlsx"),
            "Excel 파일 (*.xlsx)")
        if not path:
            return
        try:
            reporter = Reporter(
                results_dir=os.path.dirname(path) or default,
                base_dir=self.base_dir)
            out = reporter.save(self._results, self._errors,
                                filename=os.path.basename(path),
                                excluded=self._excluded)
            self._log(f"엑셀 저장 완료: {out}")
            QMessageBox.information(self, "엑셀 저장", f"저장 완료:\n{out}")
        except PermissionError:
            QMessageBox.warning(
                self, "엑셀 저장",
                "대상 파일이 열려 있어 저장하지 못했습니다.\n파일을 닫고 다시 시도하세요.")
        except Exception as e:
            self._log(f"[오류] 엑셀 저장 실패: {e}")
            QMessageBox.warning(self, "엑셀 저장", f"저장 실패: {e}")

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 다이얼로그
    def _open_settings(self):
        dlg = SettingsDialog(base_dir=self.base_dir, parent=self)
        if dlg.exec():
            self._log("설정이 변경되었습니다.")
            self._restart_scheduler()

    def _open_board_finder(self):
        if self._crawler_thread is not None and self._crawler_thread.isRunning():
            QMessageBox.information(self, "게시판 자동탐색",
                                    "수집 중에는 자동탐색을 실행할 수 없습니다.")
            return
        dlg = BoardFinderDialog(base_dir=self.base_dir,
                                sites_path=self.sites_path, parent=self)
        dlg.exec()
        self._load_sites_tree()

    def _open_board_manage(self):
        dlg = BoardManageDialog(base_dir=self.base_dir,
                                sites_path=self.sites_path, parent=self)
        dlg.exec()
        self._load_sites_tree()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 스케줄러
    def _maybe_start_scheduler(self):
        env = _read_env(os.path.join(self.base_dir, ".env"))
        if str(env.get("SCHEDULE_ENABLED", "N")).upper() != "Y":
            return
        run_time = env.get("SCHEDULE_TIME", "09:00") or "09:00"
        self._scheduler_thread = SchedulerThread(
            run_time=run_time, base_dir=self.base_dir)
        self._scheduler_thread.trigger.connect(self._on_schedule_trigger)
        self._scheduler_thread.status.connect(self._log)
        self._scheduler_thread.start()

    def _restart_scheduler(self):
        if self._scheduler_thread is not None:
            self._scheduler_thread.stop()
            self._scheduler_thread.wait(3000)
            self._scheduler_thread = None
        self._maybe_start_scheduler()

    def _on_schedule_trigger(self):
        if self._crawler_thread is not None and self._crawler_thread.isRunning():
            self._log("스케줄 트리거 무시 — 수집 진행 중")
            return
        self._log("━ 자동 스케줄 실행 ━")
        self._start_crawl(site_filter=None)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 기타
    def _log(self, text: str):
        self.log.appendPlainText(text)

    def closeEvent(self, event):
        if self._crawler_thread is not None and self._crawler_thread.isRunning():
            self._crawler_thread.stop()
            self._crawler_thread.wait(8000)
        if self._scheduler_thread is not None and self._scheduler_thread.isRunning():
            self._scheduler_thread.stop()
            self._scheduler_thread.wait(3000)
        super().closeEvent(event)
