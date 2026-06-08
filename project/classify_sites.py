# -*- coding: utf-8 -*-
"""
classify_sites.py — 크롤링 가능성 분류 → 활성화/색상 반영

진단결과(sites_diagnosis.xlsx)와 비고를 근거로 각 게시판을 3분류:
  - 가능 : 자동 크롤링 가능        → 활성화 Y, 색 없음(흰색)
  - 수동 : 사람이 직접 검색/조작 필요 → 활성화 N, 노랑
  - 불가 : 로그인必/URL無/접속불가   → 활성화 N, 빨강

적용 대상:
  1) config/sites_final.xlsx (9컬럼 마스터) — 활성화 갱신 + 행 색칠
  2) config/sites.xlsx       (7컬럼 운영)   — sites_final에서 재생성 + 행 색칠
크롤러는 활성화=Y만 수집하므로, 결과적으로 '크롤링 가능' 사이트만 수집된다.
"""
import os, shutil
import openpyxl
from openpyxl.styles import PatternFill

BASE = os.path.dirname(os.path.abspath(__file__))
FINAL = os.path.join(BASE, "config", "sites_final.xlsx")
DIAG = os.path.join(BASE, "config", "sites_diagnosis.xlsx")
OPER = os.path.join(BASE, "config", "sites.xlsx")
OPER_BAK = os.path.join(BASE, "config", "sites.backup.xlsx")
FINAL_BAK = os.path.join(BASE, "config", "sites_final.prev.xlsx")

# sites_final 컬럼(1-base): 1No 2기관명 3상태 4게시판명 5게시판URL 6페이지파라미터 7활성화 8타입 9비고
F_ORG, F_STAT, F_BOARD, F_URL, F_PAGE, F_ACTIVE, F_TYPE, F_NOTE = 2, 3, 4, 5, 6, 7, 8, 9

FILL_BAD = PatternFill("solid", fgColor="FFC7CE")   # 빨강 — 불가
FILL_MAN = PatternFill("solid", fgColor="FFEB9C")   # 노랑 — 수동검색
FILL_OK = PatternFill(fill_type=None)               # 색 없음 — 가능

MANUAL_KW = ["수동", "조회버튼", "헤더메뉴", "키워드 검색", "검색 필요",
             "개별 게시판", "나라장터", "로그인필요"]
BAD_STATUS = {"URL없음", "렌더실패", "연결실패", "SSL오류", "페이지없음", "로그인필요"}


def classify(url, status, note):
    note = str(note or "")
    if not (url and str(url).strip().startswith("http")):
        return "불가"
    if status in BAD_STATUS:
        return "불가"
    # 로그인우회로 공개(직접) URL을 이미 확보한 건은 매번 수동조작 불필요 → 정상
    if "로그인우회" in note or "공개게시판 URL 확보" in note:
        return "가능"
    if status == "진입동작필요" or any(k in note for k in MANUAL_KW):
        return "수동"
    return "가능"


def load_diag():
    d = {}
    if os.path.exists(DIAG):
        ws = openpyxl.load_workbook(DIAG).active
        for r in ws.iter_rows(min_row=2, values_only=True):
            d[r[3]] = r[4]  # url -> status
    return d


def main():
    diag = load_diag()
    shutil.copy2(FINAL, FINAL_BAK)

    # ---- 1) sites_final 활성화 + 색상 ----
    wb = openpyxl.load_workbook(FINAL)
    ws = wb.active
    last_org = ""
    counts = {"가능": 0, "수동": 0, "불가": 0}
    cat_by_row = {}
    for row in range(2, ws.max_row + 1):
        org = ws.cell(row, F_ORG).value
        if org:
            last_org = org
        url = ws.cell(row, F_URL).value
        note = ws.cell(row, F_NOTE).value
        status = diag.get(url, "(미진단)")
        cat = classify(url, status, note)
        counts[cat] += 1
        cat_by_row[row] = cat

        # 활성화 설정
        ws.cell(row, F_ACTIVE).value = "Y" if cat == "가능" else "N"
        # 행 전체 색칠 (1~9열)
        fill = FILL_OK if cat == "가능" else (FILL_MAN if cat == "수동" else FILL_BAD)
        for col in range(1, 10):
            ws.cell(row, col).fill = fill
    wb.save(FINAL)

    # ---- 2) sites.xlsx 재생성(7컬럼) + 색상 ----
    if os.path.exists(OPER):
        shutil.copy2(OPER, OPER_BAK)
    out = openpyxl.Workbook()
    ows = out.active
    ows.title = "sites"
    ows.append(["기관명", "사이트타입", "게시판명", "게시판URL",
                "페이지파라미터", "활성화", "비고"])
    last_org = ""
    orow = 1
    for row in range(2, ws.max_row + 1):
        org = ws.cell(row, F_ORG).value
        if org:
            last_org = org
        else:
            org = last_org
        url = ws.cell(row, F_URL).value
        if org is None and url is None:
            continue
        cat = cat_by_row.get(row, "가능")
        ows.append([
            org,
            ws.cell(row, F_TYPE).value or "일반",
            ws.cell(row, F_BOARD).value,
            url,
            ws.cell(row, F_PAGE).value,
            "Y" if cat == "가능" else "N",
            ws.cell(row, F_NOTE).value,
        ])
        orow += 1
        fill = FILL_OK if cat == "가능" else (FILL_MAN if cat == "수동" else FILL_BAD)
        for col in range(1, 8):
            ows.cell(orow, col).fill = fill
    out.save(OPER)

    print("분류 결과:", counts)
    print(f"  → 크롤링 진행(활성Y): {counts['가능']}개")
    print(f"  → 수동검색(노랑,N):   {counts['수동']}개")
    print(f"  → 불가(빨강,N):       {counts['불가']}개")
    print(f"저장: {FINAL}\n      {OPER}")
    print(f"백업: {FINAL_BAK}, {OPER_BAK}")


if __name__ == "__main__":
    main()
