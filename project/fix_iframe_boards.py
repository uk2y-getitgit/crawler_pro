# -*- coding: utf-8 -*-
"""
fix_iframe_boards.py — iframe(새올 eminwon 등)으로 게시판을 싣는 사이트 교정

일부 지자체는 게시판을 페이지가 아닌 <iframe src="https://eminwon...">로 싣는다.
이 경우 원래 URL을 크롤링하면 게시판이 비어 0건이 된다.
해결: 활성 사이트 중 게시판 iframe을 가진 곳을 찾아
  - 게시판URL → iframe 실제 주소로 교체
  - 타입 → '심화'(AJAX 목록 로딩이라 Playwright 필요)
sites.xlsx와 sites_final.xlsx 양쪽을 갱신한다(백업 후).
"""
import os, sys, time, shutil
import requests, urllib3
from bs4 import BeautifulSoup
from urllib.parse import urljoin
urllib3.disable_warnings()

BASE = os.path.dirname(os.path.abspath(__file__))
OPER = os.path.join(BASE, "config", "sites.xlsx")
FINAL = os.path.join(BASE, "config", "sites_final.xlsx")
import openpyxl

H = {"User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")}
# 게시판 iframe으로 인정할 호스트 키워드
IFRAME_HINT = ("eminwon", "/ofr/", "OfrNot")


def fetch(url):
    for v in (True, False):
        try:
            r = requests.get(url, headers=H, timeout=15, verify=v)
            r.encoding = r.apparent_encoding
            return r.text
        except Exception:
            time.sleep(1)
    return ""


def find_board_iframe(page_url):
    """페이지에서 게시판 iframe src(절대URL)를 찾는다. 없으면 None."""
    html = fetch(page_url)
    if not html:
        return None
    s = BeautifulSoup(html, "lxml")
    for ifr in s.find_all("iframe"):
        src = ifr.get("src") or ""
        if any(h in src for h in IFRAME_HINT):
            return urljoin(page_url, src)
    return None


def update_book(path, only_agencies):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # 컬럼 위치 자동 판별 (sites.xlsx=7col / sites_final.xlsx=9col)
    header = [str(c.value or "") for c in ws[1]]
    def col(name):
        for i, h in enumerate(header, 1):
            if name in h:
                return i
        return None
    c_org = col("기관명"); c_url = col("URL") or col("게시판URL")
    c_type = col("타입") or col("사이트타입")
    if not (c_org and c_url and c_type):
        print(f"  컬럼 판별 실패: {path}"); return 0
    last_org = ""
    changed = 0
    for r in range(2, ws.max_row + 1):
        org = ws.cell(r, c_org).value
        if org:
            last_org = org
        else:
            org = last_org
        if org not in only_agencies:
            continue
        url = str(ws.cell(r, c_url).value or "")
        if not url.startswith("http") or "eminwon" in url:
            continue
        iframe = find_board_iframe(url)
        if iframe:
            ws.cell(r, c_url).value = iframe
            ws.cell(r, c_type).value = "심화"
            changed += 1
            print(f"  ✓ {org}: → {iframe[:60]} (심화)")
    if changed:
        wb.save(path)
    return changed


def main():
    agencies = sys.argv[1:] or ["태안군", "대전 동구청"]
    print("대상 기관:", agencies)
    for path in (OPER, FINAL):
        if os.path.exists(path):
            shutil.copy2(path, path + ".ifbak")
            print(f"[{os.path.basename(path)}] 백업 후 갱신")
            n = update_book(path, set(agencies))
            print(f"  → {n}건 교체\n")


if __name__ == "__main__":
    main()
