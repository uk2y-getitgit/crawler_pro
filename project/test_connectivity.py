# -*- coding: utf-8 -*-
"""
test_connectivity.py — sites_final.xlsx 접속 상태 + 일반/심화 재판정 진단

실제 크롤링은 하지 않는다. 각 게시판 URL에 대해:
  1) 접속 가능 여부 / 실패 사유 분류
  2) 정적 HTML에 게시판 목록이 보이는지로 일반·심화 재추천

결과:
  - config/sites_diagnosis.xlsx   : 전체 진단 결과
  - config/sites_failed.xlsx      : 접속 실패/주의 사이트만 (사유 포함)
"""
import sys, os, time, re
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from requests.exceptions import (
    SSLError, ConnectTimeout, ReadTimeout, Timeout,
    ConnectionError as ReqConnError, TooManyRedirects,
)
from bs4 import BeautifulSoup
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

SRC = os.path.join(BASE_DIR, "config", "sites_final.xlsx")
OUT_ALL = os.path.join(BASE_DIR, "config", "sites_diagnosis.xlsx")
OUT_FAIL = os.path.join(BASE_DIR, "config", "sites_failed.xlsx")

HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0.0.0 Safari/537.36"),
    "Accept": ("text/html,application/xhtml+xml,application/xml;q=0.9,"
               "image/avif,image/webp,*/*;q=0.8"),
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
}

TIMEOUT = 20          # 초
SLOW_THRESHOLD = 10.0 # 이 시간(초) 넘으면 '응답느림' 경고
MAX_WORKERS = 4       # 정부서버 연결리셋 방지 위해 동시요청 낮춤
MAX_TRY = 3           # 연결오류·타임아웃 시 재시도 횟수

DATE_RE = re.compile(r"\d{4}[.\-/]\s?\d{1,2}[.\-/]\s?\d{1,2}")
LOGIN_HINT = re.compile(r"(login|signin|sso|auth|로그인|인증)", re.I)


def load_rows():
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active
    rows = []
    last_org = ""
    for r in ws.iter_rows(min_row=2, values_only=True):
        no, org, status, board, url, page_param, active, typ, note = (list(r) + [None] * 9)[:9]
        if org:
            last_org = org
        else:
            org = last_org
        rows.append({
            "no": no, "org": org, "board": board, "url": url,
            "page_param": page_param, "active": active,
            "type_now": typ, "note_old": note,
        })
    return rows


def analyze_type(html, final_url):
    """정적 HTML로 일반/심화 추천. (추천타입, 근거) 반환."""
    soup = BeautifulSoup(html, "lxml")
    for t in soup(["script", "style", "noscript"]):
        t.extract()
    text = soup.get_text(" ", strip=True)
    dates = DATE_RE.findall(html)
    rows_tr = soup.select("table tr, tbody tr")
    items_li = soup.select("ul li a, .board li, .bbs li")
    a_count = len(soup.find_all("a"))

    # SPA 루트가 비어있는 전형적 패턴
    spa_root = soup.select_one("#root, #app, [ng-app], [data-reactroot]")
    spa_empty = spa_root is not None and len(spa_root.get_text(strip=True)) < 30

    has_list = (len(rows_tr) >= 3) or (len(items_li) >= 5)
    has_dates = len(set(dates)) >= 3

    if spa_empty:
        return "심화", "SPA 루트 비어있음(JS 렌더링 필요)"
    if has_list and has_dates:
        return "일반", f"정적 목록 감지(행{len(rows_tr)}/날짜{len(set(dates))})"
    if has_dates and a_count >= 20:
        return "일반", f"정적 날짜·링크 감지(날짜{len(set(dates))})"
    if len(text) < 400 and a_count < 15:
        return "심화", "본문 거의 비어있음(JS 렌더링 추정)"
    if not has_dates and not has_list:
        return "심화", "목록·날짜 미검출(JS 렌더링 추정)"
    return "일반", f"정적 콘텐츠 존재(텍스트{len(text)}자)"


def check_deep(df, row):
    """심화 사이트: Playwright로 렌더링 후 게시판 내용 존재 여부 검증."""
    import time as _t
    url = row["url"]
    res = dict(row)
    if not (url and str(url).strip().startswith("http")):
        res.update(status="URL없음", reason="URL 비어있음/형식오류",
                   http=None, elapsed=None, type_reco="", type_reason="")
        return res
    t0 = _t.time()
    status, html, final, err = df.fetch(url)
    elapsed = round(_t.time() - t0, 1)
    if err is not None or not html:
        res.update(status="렌더실패", reason=(err or "빈 응답"),
                   http=status, elapsed=elapsed, type_reco="", type_reason="")
        return res
    reco, why = analyze_type(html, final)
    # 렌더 후에도 목록·날짜 없으면 진입동작(메뉴/검색) 필요로 판단
    if "JS 렌더링" in why or "비어있음" in why or "미검출" in why:
        res.update(status="진입동작필요",
                   reason=f"렌더 후에도 목록 미검출 — 메뉴/검색 클릭 필요({why})")
    else:
        res.update(status="정상(심화)", reason=f"브라우저 렌더링 성공({why})")
    res.update(http=status, elapsed=elapsed, type_reco=reco, type_reason=why)
    return res


def check(row):
    url = row["url"]
    res = dict(row)
    if not (url and str(url).strip().startswith("http")):
        res.update(status="URL없음", reason="URL 비어있음/형식오류",
                   http=None, elapsed=None, type_reco="", type_reason="")
        return res

    t0 = time.time()
    ssl_note = ""  # SSL 검증 실패 후 verify=False로 우회 성공 시 메모
    transient = None  # 마지막 일시오류 보관
    try:
        r = None
        # 연결리셋/타임아웃은 부하성 일시오류일 수 있어 MAX_TRY회 재시도한다.
        for attempt in range(MAX_TRY):
            try:
                r = requests.get(url, headers=HEADERS, timeout=TIMEOUT,
                                 allow_redirects=True, verify=True)
                break
            except (SSLError, ReqConnError) as e:
                # 한국 정부사이트는 GPKI 인증서 체인 문제로 SSL 검증 실패가 흔함.
                # verify=False로 재시도해 '실제 접속 가능 여부'를 확인한다.
                try:
                    r = requests.get(url, headers=HEADERS, timeout=TIMEOUT,
                                     allow_redirects=True, verify=False)
                    ssl_note = "SSL 검증 실패 → verify=False로 접속 가능(크롤러서 SSL검증 비활성화 필요)"
                    break
                except (ReqConnError, Timeout) as e2:
                    transient = e2  # 부하성 일시오류 가능 → 재시도
                except SSLError:
                    raise  # 진짜 SSL 실패
            except (ConnectTimeout, ReadTimeout, Timeout) as e:
                transient = e
            time.sleep(1.5 * (attempt + 1))  # 백오프 후 재시도
        if r is None:
            raise transient if transient else ReqConnError("연결 실패")
        elapsed = round(time.time() - t0, 1)
        code = r.status_code
        final = r.url

        # 인코딩 보정
        if not r.encoding or r.encoding.lower() == "iso-8859-1":
            r.encoding = r.apparent_encoding
        html = r.text or ""

        if code in (401, 403):
            res.update(status="로그인필요/차단", reason=f"HTTP {code} (인증 필요 또는 봇 차단)")
        elif code == 404:
            res.update(status="페이지없음", reason="HTTP 404 (URL 변경/삭제 추정)")
        elif code >= 500:
            res.update(status="서버오류", reason=f"HTTP {code}")
        elif 300 <= code < 400:
            res.update(status="리다이렉트", reason=f"HTTP {code} → {final}")
        elif LOGIN_HINT.search(final) and LOGIN_HINT.search(html[:3000]):
            res.update(status="로그인필요", reason=f"로그인 페이지로 이동: {final}")
        else:
            # 정상 응답 (verify=False 우회 포함)
            if ssl_note:
                res.update(status="SSL주의", reason=ssl_note)
            elif elapsed >= SLOW_THRESHOLD:
                res.update(status="응답느림", reason=f"정상이나 {elapsed}s 소요")
            else:
                res.update(status="정상", reason="")
            reco, why = analyze_type(html, final)
            res.update(type_reco=reco, type_reason=why)

        res.setdefault("type_reco", "")
        res.setdefault("type_reason", "")
        res.update(http=code, elapsed=elapsed)
        return res

    except SSLError:
        res.update(status="SSL오류", reason="SSL 인증서 오류")
    except (ConnectTimeout, ReadTimeout, Timeout):
        res.update(status="타임아웃", reason=f"{TIMEOUT}s 내 무응답(응답 느림/차단)")
    except TooManyRedirects:
        res.update(status="리다이렉트오류", reason="리다이렉트 무한루프")
    except ReqConnError:
        res.update(status="연결실패", reason="DNS 실패/연결 거부(URL 오류 또는 폐쇄)")
    except Exception as e:
        res.update(status="오류", reason=f"{type(e).__name__}: {e}")

    res.update(http=None, elapsed=round(time.time() - t0, 1),
               type_reco="", type_reason="")
    return res


def write_all(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "진단결과"
    head = ["No", "기관명", "게시판명", "URL", "접속상태", "HTTP", "응답(s)",
            "사유", "현재타입", "추천타입", "타입변경?", "타입근거", "활성화"]
    ws.append(head)
    for r in results:
        cur = (r.get("type_now") or "일반")
        reco = r.get("type_reco") or ""
        changed = "★" if (reco and reco != cur) else ""
        ws.append([r["no"], r["org"], r["board"], r["url"], r["status"],
                   r.get("http"), r.get("elapsed"), r.get("reason", ""),
                   cur, reco, changed, r.get("type_reason", ""), r.get("active")])
    style(ws)
    wb.save(OUT_ALL)


def write_fail(results):
    ok = {"정상", "정상(심화)"}
    bad = [r for r in results if r["status"] not in ok]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "접속실패_주의"
    ws.append(["No", "기관명", "게시판명", "URL", "접속상태", "실패사유", "HTTP", "응답(s)"])
    for r in bad:
        ws.append([r["no"], r["org"], r["board"], r["url"], r["status"],
                   r.get("reason", ""), r.get("http"), r.get("elapsed")])
    style(ws)
    wb.save(OUT_FAIL)
    return len(bad)


def style(ws):
    hfill = PatternFill("solid", fgColor="2F5496")
    hfont = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    widths = [5, 16, 20, 50, 14, 7, 8, 40, 9, 9, 9, 30, 7]
    for i, w in enumerate(widths[:ws.max_column], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def main():
    rows = load_rows()
    # 심화/일반 분리: 일반은 requests 동시처리, 심화는 Playwright 순차처리
    deep_idx = [i for i, r in enumerate(rows)
                if str(r.get("type_now") or "").strip() == "심화"]
    norm_idx = [i for i in range(len(rows)) if i not in deep_idx]
    print(f"진단 대상: {len(rows)}행 "
          f"(일반 {len(norm_idx)}: requests 동시{MAX_WORKERS} / "
          f"심화 {len(deep_idx)}: Playwright 순차)\n")
    results = [None] * len(rows)
    done = 0

    # 1) 일반 사이트 — 동시 처리
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(check, rows[i]): i for i in norm_idx}
        for f in as_completed(futs):
            i = futs[f]
            results[i] = f.result()
            done += 1
            r = results[i]
            mark = "OK " if r["status"] == "정상" else "!! "
            print(f"[{done:3}/{len(rows)}] {mark}{r['status']:12} "
                  f"{str(r['org'])[:14]:14} {str(r.get('reason',''))[:36]}")

    # 2) 심화 사이트 — Playwright 브라우저 1회 기동 후 순차 렌더링
    if deep_idx:
        from modules.playwright_crawler import DeepFetcher
        print(f"\n  --- 심화 {len(deep_idx)}건 브라우저 렌더링 ---")
        with DeepFetcher(headless=True) as df:
            for i in deep_idx:
                results[i] = check_deep(df, rows[i])
                done += 1
                r = results[i]
                mark = "OK " if r["status"].startswith("정상") else "!! "
                print(f"[{done:3}/{len(rows)}] {mark}{r['status']:12} "
                      f"{str(r['org'])[:14]:14} {str(r.get('reason',''))[:36]}")

    write_all(results)
    nfail = write_fail(results)

    from collections import Counter
    print("\n=== 접속상태 요약 ===")
    for k, v in Counter(r["status"] for r in results).most_common():
        print(f"  {k:14} {v}개")
    nchg = sum(1 for r in results
               if r.get("type_reco") and r["type_reco"] != (r.get("type_now") or "일반"))
    print(f"\n타입 변경 추천: {nchg}개")
    print(f"접속 주의/실패: {nfail}개")
    print(f"\n저장:\n  {OUT_ALL}\n  {OUT_FAIL}")


if __name__ == "__main__":
    main()
