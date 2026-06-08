# -*- coding: utf-8 -*-
"""
convert_sitelist.py — sites_final.xlsx(9컬럼) → config/sites.xlsx(7컬럼) 변환

실제 크롤러(crawler.py·gui)는 7컬럼 스키마(config/sites.xlsx)를 읽는다.
사용자가 큐레이션한 sites_final.xlsx(9컬럼)를 그 스키마로 변환해 교체한다.
기존 sites.xlsx는 sites.backup.xlsx로 백업한다.

매핑:
  sites_final[기관명]        → 기관명      (빈칸은 직전 기관명으로 채움 fill-down)
  sites_final[타입]          → 사이트타입  (None → '일반')
  sites_final[게시판명]      → 게시판명
  sites_final[게시판URL]     → 게시판URL
  sites_final[페이지파라미터] → 페이지파라미터
  sites_final[활성화]        → 활성화      (None → 'Y')
  sites_final[비고]          → 비고
"""
import os, shutil
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "config", "sites_final.xlsx")
DST = os.path.join(BASE, "config", "sites.xlsx")
BACKUP = os.path.join(BASE, "config", "sites.backup.xlsx")

# sites_final 컬럼(0-base): 0=No 1=기관명 2=상태 3=게시판명 4=게시판URL 5=페이지파라미터 6=활성화 7=타입 8=비고
HEADER7 = ["기관명", "사이트타입", "게시판명", "게시판URL", "페이지파라미터", "활성화", "비고"]


def main():
    if os.path.exists(DST):
        shutil.copy2(DST, BACKUP)
        print("기존 sites.xlsx 백업:", BACKUP)

    src = openpyxl.load_workbook(SRC).active
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "sites"
    ws.append(HEADER7)

    last_org = ""
    n, n_url, n_deep, n_active = 0, 0, 0, 0
    for r in src.iter_rows(min_row=2, values_only=True):
        row = (list(r) + [None] * 9)[:9]
        org, board, url, page, active, typ, note = (
            row[1], row[3], row[4], row[5], row[6], row[7], row[8])
        if org:
            last_org = org
        else:
            org = last_org
        if org is None and url is None:
            continue
        site_type = (typ or "일반")
        active = (active or "Y")
        ws.append([org, site_type, board, url, page, active, note])
        n += 1
        if url and str(url).strip().startswith("http"):
            n_url += 1
        if site_type == "심화":
            n_deep += 1
        if str(active).upper() == "Y":
            n_active += 1

    out.save(DST)
    print(f"변환 완료 → {DST}")
    print(f"  총 {n}행 / URL보유 {n_url} / 활성화 {n_active} / 심화 {n_deep}")


if __name__ == "__main__":
    main()
