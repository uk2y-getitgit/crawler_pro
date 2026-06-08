# -*- coding: utf-8 -*-
"""
apply_diagnosis.py — sites_diagnosis.xlsx 진단결과를 sites_final.xlsx에 반영

반영 내용:
  1) 타입 변경 추천(★) → '타입' 컬럼 갱신 (일반/심화)
     - 약한 근거(날짜<5, 행<3 등)는 비고에 '타입추정-크롤검증요' 부기
  2) 접속 주의/실패 사이트 → 타입/활성화/비고 조정
원본은 sites_final.backup.xlsx 로 백업한다.
"""
import os, sys, shutil, re
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
FINAL = os.path.join(BASE, "config", "sites_final.xlsx")
DIAG = os.path.join(BASE, "config", "sites_diagnosis.xlsx")
BACKUP = os.path.join(BASE, "config", "sites_final.backup.xlsx")

# sites_final 컬럼(1-base): No,기관명,상태,게시판명,게시판URL,페이지파라미터,활성화,타입,비고
C_NO, C_ORG, C_STAT, C_BOARD, C_URL, C_PAGE, C_ACTIVE, C_TYPE, C_NOTE = range(1, 10)

# 접속 실패/주의 사이트별 조치 (URL 일부로 매칭)
FAIL_RULES = [
    ("kospo.co.kr",      "심화", None, "SSL 핸드셰이크 실패 → 브라우저(심화) 필요"),
    ("ebiz.khnp.co.kr",  "심화", None, "로그인필요(login.do) → 공개 게시판URL 재확보 또는 헤더메뉴 내비 필요"),
    ("komipo.co.kr",     None,   None, "SSL검증 실패하나 verify=False로 접속가능"),
    ("koenergy.kr",      None,   None, "SSL검증 실패하나 verify=False로 접속가능"),
]


def add_note(ws, row, tag):
    cell = ws.cell(row=row, column=C_NOTE)
    cur = str(cell.value or "")
    if tag and tag not in cur:
        cell.value = (cur + " / " + tag).strip(" /") if cur else tag


def load_recos():
    """sites_diagnosis.xlsx에서 URL→(추천타입, 근거, 변경여부) 매핑."""
    wb = openpyxl.load_workbook(DIAG)
    ws = wb.active
    # head: No,기관명,게시판명,URL,접속상태,HTTP,응답s,사유,현재타입,추천타입,타입변경?,타입근거,활성화
    recos = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        url = r[3]
        if not url:
            continue
        recos[url] = {
            "reco": r[9], "changed": r[10] == "★", "reason": r[11] or "",
        }
    return recos


def weak_reason(reason):
    """타입근거가 약한지(경계사례) 판단."""
    m_date = re.search(r"날짜(\d+)", reason)
    m_row = re.search(r"행(\d+)", reason)
    dates = int(m_date.group(1)) if m_date else 99
    rows = int(m_row.group(1)) if m_row else 99
    if "정적 목록 감지" in reason and rows == 0 and dates < 5:
        return True
    if "정적 콘텐츠 존재" in reason:  # 텍스트 길이만 근거 → 약함
        return True
    return False


def main():
    if not os.path.exists(FINAL) or not os.path.exists(DIAG):
        print("필요 파일 없음:", FINAL, DIAG); sys.exit(1)
    shutil.copy2(FINAL, BACKUP)
    print("백업:", BACKUP)

    recos = load_recos()
    wb = openpyxl.load_workbook(FINAL)
    ws = wb.active

    n_type, n_weak, n_fail = 0, 0, 0
    for row in range(2, ws.max_row + 1):
        url = ws.cell(row=row, column=C_URL).value
        if not url:
            # URL 없는 행(정부24 등)
            add_note(ws, row, "URL 미입력 → 게시판URL 확보 필요")
            n_fail += 1
            continue

        # 1) 타입 변경 반영
        rc = recos.get(url)
        if rc and rc["changed"] and rc["reco"]:
            ws.cell(row=row, column=C_TYPE).value = rc["reco"]
            n_type += 1
            if weak_reason(rc["reason"]):
                add_note(ws, row, "타입추정-크롤검증요")
                n_weak += 1

        # 2) 접속 실패/주의 조치
        for frag, typ, active, note in FAIL_RULES:
            if frag in str(url):
                if typ:
                    ws.cell(row=row, column=C_TYPE).value = typ
                if active:
                    ws.cell(row=row, column=C_ACTIVE).value = active
                add_note(ws, row, note)
                n_fail += 1

    wb.save(FINAL)
    print(f"타입 변경 반영: {n_type}건 (약한근거 부기 {n_weak}건)")
    print(f"접속 실패/주의 조치: {n_fail}건")
    print("저장 완료:", FINAL)


if __name__ == "__main__":
    main()
